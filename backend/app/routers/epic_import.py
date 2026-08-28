"""Import from Epic's own files, with a paste-out repair loop.

    POST /api/epic-import/analyze   files (and optionally a repaired draft)
                                    -> draft + findings + a prompt to paste out
    POST /api/epic-import/diff      draft vs. an exported dataset, for checking
                                    the rules against real data

Neither endpoint writes anything and neither calls a language model. Accepting
an import happens in the wizard, through POST /api/wizard/submit, so what a user
signs off on is their rendered position rather than a file.
"""
import io
import json
import re

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from scaffold.auth import get_current_user
from scaffold.models import Grant, Price, User
from app.content_service import load_content
from app.date_utils import to_date as _to_date
from app.epic_import import (Draft, build_prompt, build_skeleton, derive_draft,
                             draft_from_payload, extract_lines, is_blocked,
                             parse_share_csv, parse_statement_lines, reconcile,
                             render_markdown, supersede_parse_findings,
                             to_wizard_payload, validate_draft)
from app.excel_io import (read_grants_from_excel, read_loans_from_excel,
                          read_prices_from_excel)

router = APIRouter(prefix="/api/epic-import", tags=["epic_import"])

_MAX_UPLOAD_BYTES = 5 * 1024 * 1024
_MAX_PASTE_BYTES = 2 * 1024 * 1024
_PDF_MAGIC = b"%PDF-"
_XLSX_MAGIC = b"PK\x03\x04"


def _read_upload(f: UploadFile | None, label: str, magic: bytes | None = None) -> bytes | None:
    if f is None or not f.filename:
        return None
    raw = f.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(raw) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail=f"{label} is too large (max 5 MB)")
    if not raw:
        return None
    if magic and not raw.startswith(magic):
        raise HTTPException(status_code=400, detail=f"{label} is not a valid file of that type")
    return raw


def _skeleton(db: Session):
    return build_skeleton(load_content(db))


def _parse_files(csv_bytes: bytes | None, pdf_bytes: bytes | None):
    """Everything readable out of the two documents, plus the text for the prompt."""
    statement, rows, findings = None, [], []
    statement_text = ""
    if pdf_bytes is not None:
        try:
            lines = extract_lines(pdf_bytes)
        except RuntimeError as e:
            raise HTTPException(status_code=503, detail=str(e))
        statement_text = "\n".join(lines)
        statement, f = parse_statement_lines(lines)
        findings += f
    csv_text = ""
    if csv_bytes is not None:
        try:
            csv_text = csv_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            csv_text = csv_bytes.decode("latin-1")
        rows, f = parse_share_csv(csv_bytes)
        findings += f
    return statement, rows, findings, statement_text, csv_text


def _payload_from_xlsx(raw: bytes) -> dict:
    """Read a filled copy of the app's own workbook into a draft payload."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not open that workbook: {e}")
    names = {s.lower(): s for s in wb.sheetnames}
    if "schedule" not in names:
        wb.close()
        raise HTTPException(status_code=400,
                            detail="That workbook has no Schedule sheet.")
    try:
        grants = read_grants_from_excel(wb[names["schedule"]])
        loans = read_loans_from_excel(wb[names["loans"]]) if "loans" in names else []
        prices = read_prices_from_excel(wb[names["prices"]]) if "prices" in names else []
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read that workbook: {e}")
    finally:
        wb.close()

    def year_of(v) -> int:
        return v.year if hasattr(v, "year") else int(v)

    by_key: dict[tuple, dict] = {}
    for g in grants:
        entry = {"year": year_of(g["year"]), "type": str(g["type"]).strip(),
                 "shares": int(g["shares"]), "price": float(g["price"]),
                 "dp_shares": int(g.get("dp_shares") or 0),
                 "election_83b": bool(g.get("election_83b")), "loans": []}
        by_key[(entry["year"], entry["type"])] = entry
    for l in loans:
        target = by_key.get((year_of(l["grant_yr"]), str(l["grant_type"]).strip()))
        if target is None:
            continue
        target["loans"].append({
            "loan_number": str(l["loan_number"] or "").strip(),
            "loan_type": str(l["loan_type"]).strip(), "loan_year": int(l["loan_year"]),
            "amount": float(l["amount"]), "interest_rate": float(l["interest_rate"]),
            "due_date": _to_date(l["due"]).isoformat(),
        })
    return {"grants": list(by_key.values()),
            "prices": [{"effective_date": _to_date(p["date"]).isoformat(),
                        "price": p["price"]} for p in prices]}


def _wizard_prefill(draft: Draft, db: Session | None = None,
                    user_id: int | None = None) -> dict:
    """The draft in the flat shape the wizard's own data loader consumes.

    Ids are negative so nothing downstream mistakes these for saved rows.

    Anything the user already has that the files do not cover is carried through
    with its real id. The wizard deletes rows its payload omits, so a draft alone
    would quietly remove price projections and grant categories the statement
    knows nothing about the moment someone accepted the import.
    """
    grants, loans, prices = [], [], []
    for gi, g in enumerate(draft.grants, 1):
        grants.append({"id": -gi, "year": g.year, "type": g.type, "shares": g.shares,
                       "price": g.price, "vest_start": g.vest_start.isoformat(),
                       "periods": g.periods,
                       "exercise_date": g.exercise_date.isoformat(),
                       "dp_shares": g.dp_shares, "election_83b": g.election_83b,
                       "version": 1})
        for li, l in enumerate(g.loans, 1):
            loans.append({"id": -(gi * 1000 + li), "grant_year": g.year,
                          "grant_type": g.type, "loan_type": l.loan_type,
                          "loan_year": l.loan_year, "amount": round(l.amount, 2),
                          "interest_rate": l.interest_rate,
                          "due_date": l.due_date.isoformat(),
                          "loan_number": l.loan_number, "refinances_loan_id": None,
                          "version": 1})
    for pi, p in enumerate(draft.prices, 1):
        prices.append({"id": -pi, "effective_date": p.effective_date.isoformat(),
                       "price": p.price, "is_estimate": False, "version": 1})

    if db is not None and user_id is not None:
        covered_years = {p.effective_date.year for p in draft.prices}
        for p in db.query(Price).filter(Price.user_id == user_id).all():
            if p.effective_date.year not in covered_years:
                prices.append({"id": p.id, "effective_date": p.effective_date.isoformat(),
                               "price": p.price, "is_estimate": bool(p.is_estimate),
                               "version": p.version or 1})
        covered_grants = {g.key for g in draft.grants}
        for g in db.query(Grant).filter(Grant.user_id == user_id).all():
            if (g.year, g.type) not in covered_grants:
                grants.append({"id": g.id, "year": g.year, "type": g.type,
                               "shares": g.shares, "price": g.price,
                               "vest_start": g.vest_start.isoformat(),
                               "periods": g.periods,
                               "exercise_date": g.exercise_date.isoformat(),
                               "dp_shares": g.dp_shares or 0,
                               "election_83b": bool(g.election_83b),
                               "version": g.version or 1})
    return {"grants": grants, "loans": loans, "prices": prices}


def _summary(draft: Draft) -> dict:
    """Figures a person can recognise, so acceptance is not a file review."""
    return {
        "grants": len(draft.grants),
        "loans": len(draft.all_loans),
        "prices": len(draft.prices),
        "total_shares": sum(g.shares for g in draft.grants),
        "total_loan_balance": round(sum(l.amount for _, l in draft.all_loans), 2),
        "grant_years": sorted({g.year for g in draft.grants}),
    }


class AnalyzeResponse(BaseModel):
    draft: dict
    wizard_payload: dict
    wizard_prefill: dict
    findings: list[dict]
    blocked: bool
    reconciles: bool
    prompt: str
    summary: dict
    origin: str


@router.post("/analyze", response_model=AnalyzeResponse)
def analyze(
    share_csv: UploadFile | None = File(default=None),
    statement_pdf: UploadFile | None = File(default=None),
    revised_draft: UploadFile | None = File(default=None),
    revised_json: str | None = Form(default=None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Read the files, or check a repaired draft against them. Writes nothing.

    Round one supplies just the files. Each later round supplies the same files
    plus whatever an assistant handed back — as pasted JSON, a .json file, or a
    filled copy of the app's workbook.
    """
    from scaffold.rate_limit import check_rate
    check_rate(user.id, "epic_import_analyze", max_calls=60, window_secs=300)

    csv_bytes = _read_upload(share_csv, "share CSV")
    pdf_bytes = _read_upload(statement_pdf, "statement PDF", _PDF_MAGIC)
    if csv_bytes is None and pdf_bytes is None:
        raise HTTPException(status_code=400,
                            detail="Upload Data for Stock Workbook, your Stock Loan "
                                   "Statement, or both")

    sk, findings = _skeleton(db)
    statement, rows, parse_findings, statement_text, csv_text = _parse_files(
        csv_bytes, pdf_bytes)
    findings += parse_findings

    payload = None
    if revised_json and revised_json.strip():
        if len(revised_json.encode()) > _MAX_PASTE_BYTES:
            raise HTTPException(status_code=400, detail="That is too much text to paste in")
        try:
            payload = json.loads(revised_json)
        except json.JSONDecodeError as e:
            raise HTTPException(
                status_code=400,
                detail=f"That is not valid JSON ({e.msg} at line {e.lineno}). Paste the "
                       f"whole object the assistant produced, starting at '{{'.")
    elif revised_draft is not None and revised_draft.filename:
        raw = _read_upload(revised_draft, "revised draft")
        if raw is not None:
            if raw.startswith(_XLSX_MAGIC):
                payload = _payload_from_xlsx(raw)
            else:
                try:
                    payload = json.loads(raw.decode("utf-8-sig"))
                except (UnicodeDecodeError, json.JSONDecodeError):
                    raise HTTPException(status_code=400,
                                        detail="That file is neither JSON nor a workbook.")

    if payload is not None:
        findings = supersede_parse_findings(findings)
        draft, f = draft_from_payload(payload, sk)
    else:
        draft, f = derive_draft(statement, rows, sk)
    findings += f
    draft.statement_date = draft.statement_date or (
        statement.statement_date if statement else None)

    # Down-payment shares are only read where rule G8 can account for them
    # exactly, so a draft carrying 0 means "could not tell", not "there were
    # none". Prefilling the wizard with that would quietly wipe a real figure
    # the moment someone accepted the import, so where the files did not say,
    # carry forward what the user already has.
    existing_dp = {(g.year, g.type): (g.dp_shares or 0)
                   for g in db.query(Grant).filter(Grant.user_id == user.id).all()}
    for g in draft.grants:
        if not g.dp_shares and existing_dp.get(g.key):
            g.dp_shares = existing_dp[g.key]

    findings += validate_draft(draft, statement, rows, sk)
    blocked = is_blocked(findings)
    reconciles = not any(x.severity == "error" for x in findings)

    return AnalyzeResponse(
        draft=draft.as_dict(),
        wizard_payload=to_wizard_payload(draft),
        wizard_prefill=_wizard_prefill(draft, db, user.id),
        findings=[x.as_dict() for x in findings],
        blocked=blocked,
        reconciles=reconciles,
        prompt=build_prompt(draft, findings, statement, sk, statement_text, csv_text),
        summary=_summary(draft),
        origin=draft.origin,
    )


# ============================================================
# DIAGNOSTICS
# ============================================================

# The dashboard's holdings report is one formatted sheet, not a dataset: two
# labelled sections with their own headers, and a share price rather than a price
# history. These are the fields it simply has no column for.
_REPORT_OMITS = frozenset({"prices", "loan_number", "periods", "vest_start",
                           "dp_shares", "election_83b"})
_GRANT_LABEL = re.compile(r"^(?P<year>\d{4})\s+(?P<type>[A-Za-z][A-Za-z -]*?)\s*$")


def _holdings_report_baseline(ws) -> tuple[list[dict], list[dict]]:
    """Read the "HOLDINGS BY GRANT" and "ACTIVE LOANS" sections of the report.

    Anchored on the section headings rather than row numbers, so adding a row to
    the title block or a column to the summary does not silently shift the read.
    """
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]

    def section(title: str) -> list[list]:
        start = next((i for i, r in enumerate(rows)
                      if str(r[0] or "").strip().upper() == title), None)
        if start is None:
            return []
        out = []
        for r in rows[start + 2:]:          # skip the heading and its column row
            label = str(r[0] or "").strip()
            if not label or label.upper() == "TOTAL":
                break
            out.append(r)
        return out

    def split(label: str) -> tuple[int, str] | None:
        m = _GRANT_LABEL.match(str(label or "").strip())
        return (int(m.group("year")), m.group("type").strip()) if m else None

    grants = []
    for r in section("HOLDINGS BY GRANT"):
        key = split(r[0])
        if key is None:
            continue
        vested, unvested = _num(r[3]), _num(r[4])
        grants.append({"year": key[0], "type": key[1],
                       "shares": int(round(vested + unvested)),
                       "price": _num(r[2]), "exercise_date": _to_date(r[1]),
                       "periods": 0, "vest_start": None,
                       "dp_shares": 0, "election_83b": False})

    loans = []
    for r in section("ACTIVE LOANS"):
        key = split(r[0])
        if key is None:
            continue
        loans.append({"loan_number": "", "grant_year": key[0], "grant_type": key[1],
                      "loan_type": str(r[1] or "").strip(), "loan_year": int(_num(r[2])),
                      "due_date": _to_date(r[3]), "amount": _num(r[4]),
                      "interest_rate": _num(r[5]), "refinances_loan_number": ""})
    return grants, loans


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace("$", "").replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _baseline_from_export(raw: bytes):
    """Read either export the app produces.

    Returns (grants, loans, prices, omits, source). The Import tab's workbook is
    the full dataset. The dashboard's holdings report is a formatted position
    statement — readable, but missing whole columns, which `omits` names so the
    reconciliation does not report the report's shape as the import's errors.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not open the export: {e}")
    names = {s.lower(): s for s in wb.sheetnames}
    if "holdings report" in names and "schedule" not in names:
        try:
            grants, loans = _holdings_report_baseline(wb[names["holdings report"]])
        except Exception as e:
            raise HTTPException(status_code=400,
                                detail=f"Could not read that holdings report: {e}")
        finally:
            wb.close()
        if not grants:
            raise HTTPException(
                status_code=400,
                detail="That holdings report has no HOLDINGS BY GRANT section to read.")
        return grants, loans, [], _REPORT_OMITS, "holdings report"
    if "schedule" not in names:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="That workbook is neither the Import tab's export nor a dashboard "
                   "holdings report — it has no Schedule sheet and no Holdings Report "
                   "sheet.")
    try:
        grants = read_grants_from_excel(wb[names["schedule"]])
        loans_raw = read_loans_from_excel(wb[names["loans"]]) if "loans" in names else []
        prices_raw = read_prices_from_excel(wb[names["prices"]]) if "prices" in names else []
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the export: {e}")
    finally:
        wb.close()

    loans = [{"loan_number": str(l["loan_number"] or "").strip(), "grant_year": l["grant_yr"],
              "grant_type": l["grant_type"], "loan_type": l["loan_type"],
              "loan_year": l["loan_year"], "amount": l["amount"],
              "interest_rate": l["interest_rate"], "due_date": _to_date(l["due"]),
              "refinances_loan_number": l.get("refinances_loan_number") or ""}
             for l in loans_raw]
    prices = [{"effective_date": _to_date(p["date"]), "price": p["price"]} for p in prices_raw]
    for g in grants:
        g["vest_start"] = _to_date(g["vest_start"])
        g["exercise_date"] = _to_date(g["exercise_date"])
    return grants, loans, prices, frozenset(), "workbook"


class DiffResponse(BaseModel):
    draft: dict
    findings: list[dict]
    report: dict
    markdown: str


@router.post("/diff", response_model=DiffResponse)
def diff(
    export_xlsx: UploadFile = File(...),
    share_csv: UploadFile | None = File(default=None),
    statement_pdf: UploadFile | None = File(default=None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Diff what the rules derive from the Epic files against exported app data.

    Reads only the files in the request — safe to run against a production export.
    """
    from scaffold.rate_limit import check_rate
    check_rate(user.id, "epic_import_diff", max_calls=20, window_secs=300)

    export_bytes = _read_upload(export_xlsx, "export xlsx", _XLSX_MAGIC)
    if export_bytes is None:
        raise HTTPException(status_code=400, detail="Upload the Excel export from the app")
    csv_bytes = _read_upload(share_csv, "share CSV")
    pdf_bytes = _read_upload(statement_pdf, "statement PDF", _PDF_MAGIC)
    if csv_bytes is None and pdf_bytes is None:
        raise HTTPException(status_code=400,
                            detail="Upload Data for Stock Workbook, your Stock Loan "
                                   "Statement, or both")

    sk, findings = _skeleton(db)
    statement, rows, parse_findings, _, _ = _parse_files(csv_bytes, pdf_bytes)
    findings += parse_findings
    draft, f = derive_draft(statement, rows, sk)
    findings += f + validate_draft(draft, statement, rows, sk)

    grants, loans, prices, omits, source = _baseline_from_export(export_bytes)
    report = reconcile(draft, grants, loans, prices, omits)
    report.counts["baseline"] = source
    return DiffResponse(draft=draft.as_dict(),
                        findings=[x.as_dict() for x in findings],
                        report=report.as_dict(),
                        markdown=render_markdown(draft, findings, report))


@router.post("/diff.md")
def diff_markdown(
    export_xlsx: UploadFile = File(...),
    share_csv: UploadFile | None = File(default=None),
    statement_pdf: UploadFile | None = File(default=None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    result = diff(export_xlsx=export_xlsx, share_csv=share_csv,
                  statement_pdf=statement_pdf, user=user, db=db)
    stamp = result.draft.get("statement_date") or "report"
    return StreamingResponse(
        io.BytesIO(result.markdown.encode("utf-8")),
        media_type="text/markdown",
        headers={"Content-Disposition": f'attachment; filename="import-diff-{stamp}.md"'},
    )
