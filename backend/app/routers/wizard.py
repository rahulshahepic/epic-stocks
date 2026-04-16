"""Wizard endpoints: tolerant structural file parsing and merge-aware bulk data save."""
import io
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel, field_validator
from sqlalchemy.orm import Session
import openpyxl

from database import get_db
from scaffold.models import User, Grant, Loan, Price, Sale, LoanPayment
from scaffold.auth import get_current_user
from app.date_utils import to_date as _to_date

router = APIRouter(prefix="/api/wizard", tags=["wizard"])


# ── Helpers ──────────────────────────────────────────────────────────────────

def _safe_int(v):
    try:
        return int(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _safe_float(v):
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _safe_date(v) -> str | None:
    if v is None:
        return None
    try:
        return _to_date(v).isoformat()
    except Exception:
        return None


# ── Parse file (tolerant — missing numbers are fine) ─────────────────────────

class ParsedGrantTemplate(BaseModel):
    year: int | None = None
    type: str | None = None
    periods: int | None = None
    vest_start: str | None = None
    exercise_date: str | None = None
    price: float | None = None


class ParsedPrice(BaseModel):
    effective_date: str
    price: float | None = None


class ParseFileResponse(BaseModel):
    grants: list[ParsedGrantTemplate]
    prices: list[ParsedPrice]


@router.post("/parse-file", response_model=ParseFileResponse)
def parse_file(
    file: UploadFile = File(...),
    _user: User = Depends(get_current_user),
):
    """Parse a structural xlsx file tolerantly — missing share counts and amounts are fine."""
    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Could not parse file as Excel (.xlsx)")

    grants: list[ParsedGrantTemplate] = []
    if "Schedule" in wb.sheetnames:
        ws = wb["Schedule"]
        for i in range(2, 100):
            yr = ws.cell(row=i, column=1).value
            if yr is None:
                break
            grants.append(ParsedGrantTemplate(
                year=_safe_int(yr),
                type=str(ws.cell(row=i, column=2).value or "").strip() or None,
                periods=_safe_int(ws.cell(row=i, column=6).value),
                vest_start=_safe_date(ws.cell(row=i, column=5).value),
                exercise_date=_safe_date(ws.cell(row=i, column=7).value),
                price=_safe_float(ws.cell(row=i, column=4).value),
            ))

    prices: list[ParsedPrice] = []
    if "Prices" in wb.sheetnames:
        ws = wb["Prices"]
        for i in range(2, 30):
            d = ws.cell(row=i, column=1).value
            if d is None:
                break
            date_str = _safe_date(d)
            if date_str:
                prices.append(ParsedPrice(
                    effective_date=date_str,
                    price=_safe_float(ws.cell(row=i, column=2).value),
                ))

    return ParseFileResponse(grants=grants, prices=prices)


# ── Shared request/response models ───────────────────────────────────────────

class WizardLoan(BaseModel):
    loan_number: str = ""
    loan_type: str  # "Purchase" or "Tax"
    loan_year: int
    amount: float
    interest_rate: float
    due_date: str
    refinances_loan_number: str = ""

    @field_validator("loan_type")
    @classmethod
    def valid_loan_type(cls, v):
        if v not in ("Purchase", "Tax", "Interest"):
            raise ValueError("loan_type must be Purchase, Tax, or Interest")
        return v

    @field_validator("amount")
    @classmethod
    def amount_positive(cls, v):
        if v <= 0:
            raise ValueError("amount must be positive")
        return v

    @field_validator("interest_rate")
    @classmethod
    def rate_non_negative(cls, v):
        if v < 0:
            raise ValueError("interest_rate cannot be negative")
        return v


class WizardGrant(BaseModel):
    year: int
    type: str  # Purchase | Catch-Up | Bonus | Free
    shares: int
    price: float
    vest_start: str
    periods: int
    exercise_date: str
    dp_shares: int = 0
    election_83b: bool = False
    loans: list[WizardLoan] = []

    @field_validator("year")
    @classmethod
    def year_range(cls, v):
        if v < 1900 or v > 2100:
            raise ValueError("year must be between 1900 and 2100")
        return v

    @field_validator("shares")
    @classmethod
    def shares_positive(cls, v):
        if v <= 0:
            raise ValueError("shares must be positive")
        return v

    @field_validator("price")
    @classmethod
    def price_non_negative(cls, v):
        if v < 0:
            raise ValueError("price cannot be negative")
        return v

    @field_validator("periods")
    @classmethod
    def periods_positive(cls, v):
        if v <= 0:
            raise ValueError("periods must be positive")
        return v


class WizardPrice(BaseModel):
    effective_date: str
    price: float

    @field_validator("price")
    @classmethod
    def price_positive(cls, v):
        if v <= 0:
            raise ValueError("price must be positive")
        return v


class WizardSubmitRequest(BaseModel):
    grants: list[WizardGrant]
    prices: list[WizardPrice]
    clear_existing: bool = True
    generate_payoff_sales: bool = True
    preserve_grant_ids: list[int] = []
    preserve_price_ids: list[int] = []


class WizardSubmitResponse(BaseModel):
    grants: int
    loans: int
    prices: int
    payoff_sales: int


# ── Preview diff models ───────────────────────────────────────────────────────

class DiffGrant(BaseModel):
    year: int
    type: str
    status: str  # "added" | "updated" | "removed" | "unchanged"
    id: int | None = None
    shares: int | None = None
    old_shares: int | None = None
    loans: int = 0
    old_loans: int | None = None


class DiffPrice(BaseModel):
    effective_date: str
    status: str  # "added" | "updated" | "removed" | "unchanged"
    id: int | None = None
    price: float | None = None
    old_price: float | None = None


class WizardPreviewResponse(BaseModel):
    grants: list[DiffGrant]
    prices: list[DiffPrice]


# ── Preview endpoint ──────────────────────────────────────────────────────────

@router.post("/preview", response_model=WizardPreviewResponse)
def preview(
    body: WizardSubmitRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Dry-run: diff wizard payload against existing DB data. No changes made."""
    existing_grants = db.query(Grant).filter(Grant.user_id == user.id).all()
    existing_prices = db.query(Price).filter(Price.user_id == user.id).all()
    existing_loans = db.query(Loan).filter(Loan.user_id == user.id).all()

    existing_grant_map = {(g.year, g.type): g for g in existing_grants}
    existing_price_map = {p.effective_date.isoformat(): p for p in existing_prices}

    # Diff grants
    grant_diffs: list[DiffGrant] = []
    wizard_grant_keys: set[tuple[int, str]] = set()
    for wg in body.grants:
        key = (wg.year, wg.type)
        wizard_grant_keys.add(key)
        existing = existing_grant_map.get(key)
        wl_count = len(wg.loans)
        if existing is None:
            grant_diffs.append(DiffGrant(
                year=wg.year, type=wg.type, status="added",
                shares=wg.shares, loans=wl_count,
            ))
        else:
            old_loans = sum(1 for l in existing_loans if l.grant_year == existing.year and l.grant_type == existing.type)
            changed = (
                existing.shares != wg.shares
                or abs(existing.price - wg.price) > 0.001
                or existing.periods != wg.periods
            )
            grant_diffs.append(DiffGrant(
                year=wg.year, type=wg.type,
                status="updated" if changed else "unchanged",
                id=existing.id, shares=wg.shares, old_shares=existing.shares,
                loans=wl_count, old_loans=old_loans,
            ))

    # Removed grants (in DB but not in wizard)
    for key, g in existing_grant_map.items():
        if key not in wizard_grant_keys:
            old_loans = sum(1 for l in existing_loans if l.grant_year == g.year and l.grant_type == g.type)
            grant_diffs.append(DiffGrant(
                year=g.year, type=g.type, status="removed",
                id=g.id, old_shares=g.shares, old_loans=old_loans,
            ))

    # Diff prices
    price_diffs: list[DiffPrice] = []
    wizard_price_keys: set[str] = set()
    for wp in body.prices:
        wizard_price_keys.add(wp.effective_date)
        existing = existing_price_map.get(wp.effective_date)
        if existing is None:
            price_diffs.append(DiffPrice(
                effective_date=wp.effective_date, status="added", price=wp.price,
            ))
        else:
            changed = abs(existing.price - wp.price) > 0.001
            price_diffs.append(DiffPrice(
                effective_date=wp.effective_date,
                status="updated" if changed else "unchanged",
                id=existing.id, price=wp.price, old_price=existing.price,
            ))

    # Removed prices
    for key, p in existing_price_map.items():
        if key not in wizard_price_keys:
            price_diffs.append(DiffPrice(
                effective_date=key, status="removed",
                id=p.id, old_price=p.price,
            ))

    return WizardPreviewResponse(grants=grant_diffs, prices=price_diffs)


# ── Submit ────────────────────────────────────────────────────────────────────

@router.post("/submit", response_model=WizardSubmitResponse, status_code=201)
def submit(
    body: WizardSubmitRequest,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Save wizard data. clear_existing=True nukes all prior data; False merges."""
    loan_objects: list[tuple[Loan, str]] = []  # (loan_obj, refinances_loan_number)

    if body.clear_existing:
        db.query(LoanPayment).filter(LoanPayment.user_id == user.id).delete()
        db.query(Sale).filter(Sale.user_id == user.id).delete()
        db.query(Grant).filter(Grant.user_id == user.id).delete()
        db.query(Loan).filter(Loan.user_id == user.id).delete()
        db.query(Price).filter(Price.user_id == user.id).delete()

        price_count = _insert_prices(body.prices, user.id, db)
        grant_count = _insert_grants(body.grants, user.id, db)
        db.flush()
        loan_count = _insert_loans(body.grants, user.id, db, loan_objects)
    else:
        price_count, grant_count, loan_count = _merge(body, user.id, db, loan_objects)

    db.flush()

    # Resolve refinance references across all user loans
    all_loans_by_number = {
        l.loan_number: l
        for l in db.query(Loan).filter(Loan.user_id == user.id).all()
        if l.loan_number
    }
    for loan, ref_num in loan_objects:
        if ref_num and ref_num in all_loans_by_number:
            loan.refinances_loan_id = all_loans_by_number[ref_num].id

    db.flush()

    # Payoff sales for all loans — skip refinanced loans (their payoff events
    # are converted to $0 "Refinanced" events, so a linked sale would be confusing)
    refinanced_ids = {loan.refinances_loan_id for loan, _ in loan_objects if loan.refinances_loan_id is not None}
    payoff_count = 0
    if body.generate_payoff_sales:
        from app.routers.loans import _compute_payoff_sale
        for loan, _ in loan_objects:
            if loan.id not in refinanced_ids:
                try:
                    suggestion = _compute_payoff_sale(loan, user, db)
                    if suggestion["shares"] > 0 and suggestion["price_per_share"] > 0:
                        db.add(Sale(
                            user_id=user.id,
                            date=suggestion["date"],
                            shares=suggestion["shares"],
                            price_per_share=suggestion["price_per_share"],
                            loan_id=loan.id,
                            notes=suggestion["notes"],
                        ))
                        payoff_count += 1
                except Exception:
                    pass  # Don't abort the whole wizard if payoff calc fails

    db.commit()

    from app.event_cache import schedule_recompute
    schedule_recompute(user.id)

    return WizardSubmitResponse(
        grants=grant_count,
        loans=loan_count,
        prices=price_count,
        payoff_sales=payoff_count,
    )


# ── Internal helpers ──────────────────────────────────────────────────────────

def _insert_prices(prices: list[WizardPrice], user_id: int, db: Session) -> int:
    count = 0
    for p in prices:
        db.add(Price(
            user_id=user_id,
            effective_date=_to_date(p.effective_date),
            price=p.price,
            is_estimate=_to_date(p.effective_date) > date.today(),
        ))
        count += 1
    return count


def _insert_grants(grants: list[WizardGrant], user_id: int, db: Session) -> int:
    count = 0
    for g in grants:
        db.add(Grant(
            user_id=user_id,
            year=g.year, type=g.type, shares=g.shares, price=g.price,
            vest_start=_to_date(g.vest_start), periods=g.periods,
            exercise_date=_to_date(g.exercise_date),
            dp_shares=g.dp_shares, election_83b=g.election_83b,
        ))
        count += 1
    return count


def _insert_loans(
    grants: list[WizardGrant],
    user_id: int,
    db: Session,
    loan_objects: list,
) -> int:
    count = 0
    for g in grants:
        for wl in g.loans:
            loan = Loan(
                user_id=user_id,
                grant_year=g.year, grant_type=g.type,
                loan_type=wl.loan_type, loan_year=wl.loan_year,
                amount=wl.amount, interest_rate=wl.interest_rate,
                due_date=_to_date(wl.due_date),
                loan_number=wl.loan_number or None,
            )
            db.add(loan)
            loan_objects.append((loan, wl.refinances_loan_number))
            count += 1
    return count


def _merge(
    body: WizardSubmitRequest,
    user_id: int,
    db: Session,
    loan_objects: list,
) -> tuple[int, int, int]:
    """Upsert grants/prices/loans; delete orphans not in preserve lists."""
    existing_grants = {
        (g.year, g.type): g
        for g in db.query(Grant).filter(Grant.user_id == user_id).all()
    }
    existing_prices = {
        p.effective_date.isoformat(): p
        for p in db.query(Price).filter(Price.user_id == user_id).all()
    }

    wizard_grant_keys = {(g.year, g.type) for g in body.grants}
    wizard_price_keys = {p.effective_date for p in body.prices}
    preserve_grant_ids = set(body.preserve_grant_ids)
    preserve_price_ids = set(body.preserve_price_ids)

    # Delete auto-generated payoff sales — they'll be regenerated from the new loan set
    db.query(Sale).filter(Sale.user_id == user_id, Sale.loan_id.isnot(None)).delete()

    # Delete orphaned grants (not in wizard, not preserved) and their loans
    for key, g in existing_grants.items():
        if key not in wizard_grant_keys and g.id not in preserve_grant_ids:
            db.query(Loan).filter(
                Loan.user_id == user_id,
                Loan.grant_year == g.year,
                Loan.grant_type == g.type,
            ).delete()
            db.delete(g)

    # Delete orphaned prices
    for key, p in existing_prices.items():
        if key not in wizard_price_keys and p.id not in preserve_price_ids:
            db.delete(p)

    db.flush()

    # Upsert prices
    price_count = 0
    for wp in body.prices:
        existing = existing_prices.get(wp.effective_date)
        if existing:
            existing.price = wp.price
            existing.is_estimate = _to_date(wp.effective_date) > date.today()
        else:
            db.add(Price(
                user_id=user_id,
                effective_date=_to_date(wp.effective_date),
                price=wp.price,
                is_estimate=_to_date(wp.effective_date) > date.today(),
            ))
        price_count += 1

    # Upsert grants
    grant_count = 0
    for wg in body.grants:
        existing = existing_grants.get((wg.year, wg.type))
        if existing:
            existing.shares = wg.shares
            existing.price = wg.price
            existing.vest_start = _to_date(wg.vest_start)
            existing.periods = wg.periods
            existing.exercise_date = _to_date(wg.exercise_date)
            existing.dp_shares = wg.dp_shares
            existing.election_83b = wg.election_83b
        else:
            db.add(Grant(
                user_id=user_id, year=wg.year, type=wg.type,
                shares=wg.shares, price=wg.price,
                vest_start=_to_date(wg.vest_start), periods=wg.periods,
                exercise_date=_to_date(wg.exercise_date),
                dp_shares=wg.dp_shares, election_83b=wg.election_83b,
            ))
        grant_count += 1

    db.flush()

    # Upsert loans per grant; delete orphans within each grant's loan set
    loan_count = 0
    for wg in body.grants:
        existing_loans_for_grant = db.query(Loan).filter(
            Loan.user_id == user_id,
            Loan.grant_year == wg.year,
            Loan.grant_type == wg.type,
        ).all()

        processed_loan_ids: set[int] = set()

        for wl in wg.loans:
            # Match by loan_number if present, otherwise by (loan_type, loan_year)
            if wl.loan_number:
                match = next((l for l in existing_loans_for_grant if l.loan_number == wl.loan_number), None)
            else:
                match = next(
                    (l for l in existing_loans_for_grant if l.loan_type == wl.loan_type and l.loan_year == wl.loan_year),
                    None,
                )

            if match:
                match.loan_type = wl.loan_type
                match.loan_year = wl.loan_year
                match.amount = wl.amount
                match.interest_rate = wl.interest_rate
                match.due_date = _to_date(wl.due_date)
                if wl.loan_number:
                    match.loan_number = wl.loan_number
                processed_loan_ids.add(match.id)
                loan_objects.append((match, wl.refinances_loan_number))
            else:
                new_loan = Loan(
                    user_id=user_id,
                    grant_year=wg.year, grant_type=wg.type,
                    loan_type=wl.loan_type, loan_year=wl.loan_year,
                    amount=wl.amount, interest_rate=wl.interest_rate,
                    due_date=_to_date(wl.due_date),
                    loan_number=wl.loan_number or None,
                )
                db.add(new_loan)
                loan_objects.append((new_loan, wl.refinances_loan_number))
            loan_count += 1

        # Delete loans for this grant that weren't in the wizard
        for l in existing_loans_for_grant:
            if l.id not in processed_loan_ids:
                db.delete(l)

    return price_count, grant_count, loan_count
