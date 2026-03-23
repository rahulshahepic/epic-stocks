"""Import/export endpoints for Excel files."""
import io
import json
import tempfile
from datetime import datetime, date
from openpyxl.comments import Comment
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from database import get_db
from models import User, Grant, Loan, Price, Sale, ImportBackup
from auth import get_current_user
from excel_io import read_grants_from_excel, read_prices_from_excel, read_loans_from_excel, write_events_to_excel
from core import generate_all_events, compute_timeline
from schemas import LOAN_TYPES

import openpyxl
from openpyxl.styles import Font, PatternFill

_MAX_BACKUPS_PER_USER = 3

router = APIRouter(prefix="/api", tags=["import_export"])


def _to_date(val) -> date:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return datetime.strptime(str(val), "%Y-%m-%d").date()


def _to_year(val) -> int:
    """Convert a value that might be a date or int to a year integer."""
    if isinstance(val, (datetime, date)):
        return val.year
    return int(val)


# ============================================================
# VALIDATION HELPERS
# ============================================================

def _validate_grant(g: dict, row: int) -> list[str]:
    errors = []
    year = g.get("year")
    if not isinstance(year, (int, float)) or int(year) < 1900 or int(year) > 2100:
        errors.append(f"Row {row}: year must be between 1900 and 2100")
    gtype = g.get("type")
    if not gtype or not str(gtype).strip():
        errors.append(f"Row {row}: type cannot be empty")
    shares = g.get("shares")
    if not isinstance(shares, (int, float)) or int(shares) <= 0:
        errors.append(f"Row {row}: shares must be positive")
    price = g.get("price")
    if not isinstance(price, (int, float)) or float(price) < 0:
        errors.append(f"Row {row}: price cannot be negative")
    periods = g.get("periods")
    if not isinstance(periods, (int, float)) or int(periods) <= 0:
        errors.append(f"Row {row}: periods must be positive")
    # dp_shares can be negative (DP exchange returns shares)
    for field in ("vest_start", "exercise_date"):
        v = g.get(field)
        if v is None:
            errors.append(f"Row {row}: {field} is required")
        else:
            try:
                _to_date(v)
            except Exception:
                errors.append(f"Row {row}: {field} is not a valid date")
    return errors


def _validate_loan(ln: dict, row: int) -> list[str]:
    errors = []
    for yr_field in ("grant_yr", "loan_year"):
        v = ln.get(yr_field)
        try:
            yr = _to_year(v) if v is not None else None
            if yr is None or yr < 1900 or yr > 2100:
                errors.append(f"Row {row}: {yr_field} must be between 1900 and 2100")
        except (ValueError, TypeError):
            errors.append(f"Row {row}: {yr_field} is not a valid year")
    gtype = ln.get("grant_type")
    if not gtype or not str(gtype).strip():
        errors.append(f"Row {row}: grant_type cannot be empty")
    ltype = (ln.get("loan_type") or "").strip()
    if ltype not in LOAN_TYPES:
        errors.append(f"Row {row}: loan_type must be one of {sorted(LOAN_TYPES)}, got '{ltype}'")
    amt = ln.get("amount")
    if not isinstance(amt, (int, float)) or float(amt) <= 0:
        errors.append(f"Row {row}: amount must be positive")
    rate = ln.get("interest_rate")
    if not isinstance(rate, (int, float)) or float(rate) < 0:
        errors.append(f"Row {row}: interest_rate cannot be negative")
    due = ln.get("due")
    if due is None:
        errors.append(f"Row {row}: due_date is required")
    else:
        try:
            _to_date(due)
        except Exception:
            errors.append(f"Row {row}: due_date is not a valid date")
    return errors


def _validate_price(p: dict, row: int) -> list[str]:
    errors = []
    d = p.get("date")
    if d is None:
        errors.append(f"Row {row}: date is required")
    else:
        try:
            _to_date(d)
        except Exception:
            errors.append(f"Row {row}: date is not a valid date")
    price = p.get("price")
    if not isinstance(price, (int, float)) or float(price) <= 0:
        errors.append(f"Row {row}: price must be positive")
    return errors


# ============================================================
# IMPORT (now supports partial — only sheets that exist)
# ============================================================

_MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB
_XLSX_MAGIC = b"PK\x03\x04"  # ZIP/OOXML magic bytes


@router.post("/import/excel", status_code=201)
def import_excel(
    file: UploadFile = File(...),
    generate_payoff_sales: bool = Query(default=False),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel (.xlsx) file")

    raw = file.file.read(_MAX_UPLOAD_BYTES + 1)
    if len(raw) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=400, detail="File too large (max 5 MB)")
    if not raw.startswith(_XLSX_MAGIC):
        raise HTTPException(status_code=400, detail="File is not a valid Excel (.xlsx) file")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=True) as tmp:
        tmp.write(raw)
        tmp.flush()
        try:
            wb = openpyxl.load_workbook(tmp.name)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Failed to open Excel file: {e}")

    sheet_names = [s.lower() for s in wb.sheetnames]

    # Read whichever sheets exist
    grants_raw = []
    prices_raw = []
    loans_raw = []
    has_schedule = "schedule" in sheet_names
    has_prices = "prices" in sheet_names
    has_loans = "loans" in sheet_names

    if not has_schedule and not has_prices and not has_loans:
        wb.close()
        raise HTTPException(
            status_code=400,
            detail="No recognized sheets found. Expected one or more of: Schedule, Prices, Loans"
        )

    try:
        if has_schedule:
            ws_name = wb.sheetnames[[s.lower() for s in wb.sheetnames].index("schedule")]
            grants_raw = read_grants_from_excel(wb[ws_name])
        if has_prices:
            ws_name = wb.sheetnames[[s.lower() for s in wb.sheetnames].index("prices")]
            prices_raw = read_prices_from_excel(wb[ws_name])
        if has_loans:
            ws_name = wb.sheetnames[[s.lower() for s in wb.sheetnames].index("loans")]
            loans_raw = read_loans_from_excel(wb[ws_name])
    except Exception as e:
        wb.close()
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")
    finally:
        wb.close()

    # Validate all rows, collect errors
    all_errors = []
    for i, g in enumerate(grants_raw):
        all_errors.extend(_validate_grant(g, i + 2))
    for i, p in enumerate(prices_raw):
        all_errors.extend(_validate_price(p, i + 2))
    for i, ln in enumerate(loans_raw):
        all_errors.extend(_validate_loan(ln, i + 2))

    # Check for duplicate (year, type) within the imported grants
    seen_grants: set[tuple] = set()
    for i, g in enumerate(grants_raw):
        key = (_to_year(g["year"]), str(g.get("type", "")).strip())
        if key in seen_grants:
            all_errors.append(f"Duplicate grant: {key[1]} {key[0]} appears more than once in the Schedule sheet")
        seen_grants.add(key)

    if all_errors:
        raise HTTPException(status_code=400, detail="Validation errors:\n" + "\n".join(all_errors))

    # Snapshot current data before wiping (keep last 3 backups per user)
    _save_import_backup(user.id, has_schedule, has_prices, has_loans, db)

    # Only wipe data types that were in the uploaded file
    if has_schedule:
        db.query(Grant).filter(Grant.user_id == user.id).delete()
    if has_loans:
        # Remove loan-linked payoff sales first to avoid orphaned loan_id references
        db.query(Sale).filter(Sale.user_id == user.id, Sale.loan_id.isnot(None)).delete()
        db.query(Loan).filter(Loan.user_id == user.id).delete()
    if has_prices:
        db.query(Price).filter(Price.user_id == user.id).delete()

    for g in grants_raw:
        db.add(Grant(
            user_id=user.id,
            year=g["year"],
            type=g["type"],
            shares=g["shares"],
            price=g["price"],
            vest_start=_to_date(g["vest_start"]),
            periods=g["periods"],
            exercise_date=_to_date(g["exercise_date"]),
            dp_shares=g.get("dp_shares", 0),
        ))

    for p in prices_raw:
        db.add(Price(
            user_id=user.id,
            effective_date=_to_date(p["date"]),
            price=p["price"],
        ))

    for ln in loans_raw:
        db.add(Loan(
            user_id=user.id,
            grant_year=_to_year(ln["grant_yr"]),
            grant_type=ln["grant_type"],
            loan_type=ln["loan_type"].strip(),
            loan_year=_to_year(ln["loan_year"]),
            amount=ln["amount"],
            interest_rate=ln["interest_rate"],
            due_date=_to_date(ln["due"]),
            loan_number=str(ln.get("loan_number") or ""),
        ))

    db.commit()

    payoff_sales_created = 0
    if has_loans and generate_payoff_sales:
        from routers.loans import _compute_payoff_sale
        new_loans = db.query(Loan).filter(Loan.user_id == user.id).all()
        for ln in new_loans:
            try:
                suggestion = _compute_payoff_sale(ln, user, db)
                if suggestion["shares"] > 0 and suggestion["price_per_share"] > 0:
                    db.add(Sale(
                        user_id=user.id,
                        date=suggestion["date"],
                        shares=suggestion["shares"],
                        price_per_share=suggestion["price_per_share"],
                        loan_id=ln.id,
                        notes=suggestion["notes"],
                    ))
                    payoff_sales_created += 1
            except Exception:
                pass  # best-effort; missing price data etc. silently skipped
        if payoff_sales_created:
            db.commit()

    return {
        "grants": len(grants_raw),
        "prices": len(prices_raw),
        "loans": len(loans_raw),
        "payoff_sales": payoff_sales_created,
        "sheets_imported": [s for s, present in [("Schedule", has_schedule), ("Prices", has_prices), ("Loans", has_loans)] if present],
    }


# ============================================================
# BACKUP HELPERS & ENDPOINTS
# ============================================================

def _save_import_backup(user_id: int, has_schedule: bool, has_prices: bool, has_loans: bool, db: Session):
    """Snapshot the affected data types before an import overwrites them."""
    grants = []
    prices = []
    loans = []
    if has_schedule:
        for g in db.query(Grant).filter(Grant.user_id == user_id).all():
            grants.append({
                "year": g.year, "type": g.type, "shares": g.shares, "price": g.price,
                "vest_start": str(g.vest_start), "periods": g.periods,
                "exercise_date": str(g.exercise_date), "dp_shares": g.dp_shares or 0,
            })
    if has_prices:
        for p in db.query(Price).filter(Price.user_id == user_id).all():
            prices.append({"effective_date": str(p.effective_date), "price": p.price})
    if has_loans:
        for ln in db.query(Loan).filter(Loan.user_id == user_id).all():
            loans.append({
                "grant_year": ln.grant_year, "grant_type": ln.grant_type,
                "loan_type": ln.loan_type, "loan_year": ln.loan_year,
                "amount": ln.amount, "interest_rate": ln.interest_rate,
                "due_date": str(ln.due_date), "loan_number": ln.loan_number or "",
            })

    if not grants and not prices and not loans:
        return  # Nothing to back up

    db.add(ImportBackup(user_id=user_id, data_json=json.dumps(
        {"grants": grants, "prices": prices, "loans": loans}
    )))
    db.flush()  # ensure new backup is visible in the trimming query
    # Trim to last _MAX_BACKUPS_PER_USER
    all_backups = (
        db.query(ImportBackup)
        .filter(ImportBackup.user_id == user_id)
        .order_by(ImportBackup.created_at.desc())
        .all()
    )
    for old in all_backups[_MAX_BACKUPS_PER_USER:]:
        db.delete(old)


class BackupOut(BaseModel):
    id: int
    created_at: str
    grant_count: int
    price_count: int
    loan_count: int


@router.get("/import/backups", response_model=list[BackupOut])
def list_import_backups(user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    backups = (
        db.query(ImportBackup)
        .filter(ImportBackup.user_id == user.id)
        .order_by(ImportBackup.created_at.desc())
        .all()
    )
    result = []
    for b in backups:
        data = json.loads(b.data_json)
        result.append(BackupOut(
            id=b.id,
            created_at=b.created_at.isoformat(),
            grant_count=len(data.get("grants", [])),
            price_count=len(data.get("prices", [])),
            loan_count=len(data.get("loans", [])),
        ))
    return result


@router.post("/import/backups/{backup_id}/restore", status_code=200)
def restore_import_backup(
    backup_id: int,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    backup = db.query(ImportBackup).filter(
        ImportBackup.id == backup_id, ImportBackup.user_id == user.id
    ).first()
    if not backup:
        raise HTTPException(status_code=404, detail="Backup not found")

    data = json.loads(backup.data_json)
    grants = data.get("grants", [])
    prices = data.get("prices", [])
    loans = data.get("loans", [])

    if grants:
        db.query(Grant).filter(Grant.user_id == user.id).delete()
    if prices:
        db.query(Price).filter(Price.user_id == user.id).delete()
    if loans:
        db.query(Sale).filter(Sale.user_id == user.id, Sale.loan_id.isnot(None)).delete()
        db.query(Loan).filter(Loan.user_id == user.id).delete()

    for g in grants:
        db.add(Grant(
            user_id=user.id, year=g["year"], type=g["type"],
            shares=g["shares"], price=g["price"],
            vest_start=_to_date(g["vest_start"]), periods=g["periods"],
            exercise_date=_to_date(g["exercise_date"]), dp_shares=g.get("dp_shares", 0),
        ))
    for p in prices:
        db.add(Price(user_id=user.id, effective_date=_to_date(p["effective_date"]), price=p["price"]))
    for ln in loans:
        db.add(Loan(
            user_id=user.id, grant_year=ln["grant_year"], grant_type=ln["grant_type"],
            loan_type=ln["loan_type"], loan_year=ln["loan_year"],
            amount=ln["amount"], interest_rate=ln["interest_rate"],
            due_date=_to_date(ln["due_date"]), loan_number=ln.get("loan_number", ""),
        ))

    db.commit()
    return {
        "restored_grants": len(grants),
        "restored_prices": len(prices),
        "restored_loans": len(loans),
    }


# ============================================================
# TEMPLATE DOWNLOAD
# ============================================================

@router.get("/import/template")
def download_template():
    """Download an empty Excel template with correct headers and example rows."""
    wb = openpyxl.Workbook()

    # Schedule sheet
    ws = wb.active
    ws.title = "Schedule"
    sched_headers = ["Year", "Type", "Shares", "Price", "Vest Start", "Periods",
                     "Exercise Date", "DP Shares"]
    _write_headers(ws, sched_headers)
    # Example row
    _body_cell(ws, 2, 1, 2020)
    _body_cell(ws, 2, 2, "Purchase")
    _body_cell(ws, 2, 3, 10000)
    _body_cell(ws, 2, 4, 5.00, "\\$#,##0.00")
    _body_cell(ws, 2, 5, date(2020, 3, 15), "mm/dd/yyyy")
    _body_cell(ws, 2, 6, 5)
    _body_cell(ws, 2, 7, date(2030, 3, 15), "mm/dd/yyyy")
    _body_cell(ws, 2, 8, 0)
    # Add hints as cell comments on headers (avoids extra rows that break import)
    for col, note in [(1, "e.g. 2020"), (2, "Purchase or Bonus"), (3, "# of shares"),
                      (4, "$ per share"), (5, "mm/dd/yyyy"), (6, "# vesting periods")]:
        ws.cell(row=1, column=col).comment = Comment(note, "Template")

    # Loans sheet
    ws_loans = wb.create_sheet("Loans")
    loan_headers = ["Loan #", "Grant Year", "Grant Type", "Loan Type", "Loan Year",
                    "Amount", "Rate", "Due Date"]
    _write_headers(ws_loans, loan_headers)
    _body_cell(ws_loans, 2, 1, "L001")
    _body_cell(ws_loans, 2, 2, 2020)
    _body_cell(ws_loans, 2, 3, "Purchase")
    _body_cell(ws_loans, 2, 4, "Interest")
    _body_cell(ws_loans, 2, 5, 2020)
    _body_cell(ws_loans, 2, 6, 5000.00, "\\$#,##0.00")
    _body_cell(ws_loans, 2, 7, 0.05, "0.00%")
    _body_cell(ws_loans, 2, 8, date(2025, 3, 15), "mm/dd/yyyy")
    for col, note in [(3, "Purchase or Bonus"), (4, "Interest, Tax, Principal, or Purchase"),
                      (7, "decimal, e.g. 0.05 = 5%")]:
        ws_loans.cell(row=1, column=col).comment = Comment(note, "Template")

    # Prices sheet
    ws_prices = wb.create_sheet("Prices")
    _write_headers(ws_prices, ["Date", "Price"])
    _body_cell(ws_prices, 2, 1, date(2020, 1, 1), "mm/dd/yyyy")
    _body_cell(ws_prices, 2, 2, 5.00, "\\$#,##0.00")
    ws_prices.cell(row=1, column=1).comment = Comment("One row per annual price update", "Template")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Vesting_Template.xlsx"},
    )


# ============================================================
# SAMPLE DOWNLOAD
# ============================================================

@router.get("/import/sample")
def download_sample():
    """Download a sample Excel file pre-filled with realistic fake data and explanatory comments."""
    from openpyxl.comments import Comment

    wb = openpyxl.Workbook()

    # ---- Schedule sheet ----
    ws = wb.active
    ws.title = "Schedule"
    _write_headers(ws, ["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"])

    # Header comments
    header_comments = [
        (1, "The year of this grant — matches your Epic annual statement (e.g. 2020)."),
        (2, "Grant type: Purchase, Bonus, Catch-Up, Free, etc."),
        (3, "Total shares in this grant."),
        (4, "Purchase price per share ($0 for RSU/Bonus grants, strike price for options)."),
        (5, "Date vesting begins (the day Epic starts the vesting clock)."),
        (6, "Number of vesting periods (e.g. 8 = quarterly over 4 years, 4 = quarterly over 2 years)."),
        (7, "Date by which options must be exercised. For RSU/Bonus grants use Dec 31 of the grant year."),
        (8, "Down-payment shares used to buy this grant (from purchase confirmation). "
            "Negative means Epic returned shares to you. 0 if none."),
    ]
    for col, note in header_comments:
        ws.cell(row=1, column=col).comment = Comment(note, "Sample")

    sched_rows = [
        (2020, "Purchase", 50000, 1.85, date(2022, 6, 15), 5, date(2020, 12, 31), 0),
        (2020, "Bonus",    5000,  0.00, date(2022, 6, 15), 4, date(2020, 12, 31), 0),
        (2021, "Purchase", 80000, 2.20, date(2023, 9, 30), 5, date(2021, 12, 31), 0),
        (2022, "Purchase", 100000, 2.75, date(2024, 9, 30), 4, date(2022, 12, 31), 0),
        (2022, "Bonus",    20000, 0.00, date(2027, 9, 30), 1, date(2022, 12, 31), 0),
        (2023, "Purchase", 120000, 3.10, date(2025, 9, 30), 4, date(2023, 12, 31), 0),
        (2024, "Purchase", 150000, 3.60, date(2026, 9, 30), 4, date(2024, 12, 31), -15000),
    ]
    fmts = [None, None, "#,##0", "\\$#,##0.00", "mm/dd/yyyy", None, "mm/dd/yyyy", "#,##0"]
    for r, row_data in enumerate(sched_rows, 2):
        for c, (val, fmt) in enumerate(zip(row_data, fmts), 1):
            _body_cell(ws, r, c, val, fmt)

    # Annotate first data row
    first_row_notes = [
        (1, "2020 = grant issued in 2020"),
        (2, "Purchase = standard yearly purchase grant"),
        (3, "50,000 total shares in this grant"),
        (4, "$1.85 = price per share at purchase (from statement)"),
        (5, "Vesting starts June 15, 2022"),
        (6, "5 periods = 5 quarterly vests"),
        (7, "Exercise deadline Dec 31, 2020 (RSU-style — use grant year end)"),
        (8, "0 = no down-payment shares used"),
    ]
    for col, note in first_row_notes:
        ws.cell(row=2, column=col).comment = Comment(note, "Sample")

    # DP Shares note on the row that uses it
    ws.cell(row=len(sched_rows) + 1, column=8).comment = Comment(
        "-15,000 = Epic returned 15,000 shares as a down-payment credit (enter as negative)", "Sample"
    )

    # ---- Loans sheet ----
    ws_loans = wb.create_sheet("Loans")
    _write_headers(ws_loans, ["Loan #", "Grant Year", "Grant Type", "Loan Type", "Loan Year",
                               "Amount", "Rate", "Due Date"])

    loan_header_comments = [
        (1, "Loan number from your Epic statement (e.g. 002001). Used to match payoff sales."),
        (2, "Year of the grant this loan is associated with."),
        (3, "Grant type — must match the Schedule sheet (e.g. Purchase, Bonus)."),
        (4, "Loan type: Purchase (original purchase loan), Interest (accrued interest loan), "
            "Tax (tax withholding loan)."),
        (5, "Year this loan was issued."),
        (6, "Principal amount of the loan in dollars."),
        (7, "Annual interest rate as a decimal (e.g. 0.0095 = 0.95%)."),
        (8, "Loan due date — when you must repay it."),
    ]
    for col, note in loan_header_comments:
        ws_loans.cell(row=1, column=col).comment = Comment(note, "Sample")

    loan_rows = [
        ("002001", 2020, "Purchase", "Purchase",  2020, 92500.00,  0.0095, date(2029, 7, 15)),
        ("002002", 2020, "Purchase", "Interest",  2022,   850.00,  0.0200, date(2029, 7, 15)),
        ("002003", 2020, "Purchase", "Interest",  2023,  1100.00,  0.0350, date(2029, 7, 15)),
        ("003001", 2021, "Purchase", "Purchase",  2021, 176000.00, 0.0095, date(2030, 7, 15)),
        ("003002", 2021, "Purchase", "Interest",  2023,  1620.00,  0.0350, date(2030, 7, 15)),
        ("004001", 2022, "Purchase", "Purchase",  2022, 275000.00, 0.0200, date(2031, 6, 30)),
        ("005001", 2023, "Purchase", "Purchase",  2023, 372000.00, 0.0350, date(2032, 6, 30)),
        ("006001", 2024, "Purchase", "Purchase",  2024, 540000.00, 0.0400, date(2033, 6, 30)),
    ]
    loan_fmts = [None, None, None, None, None, "\\$#,##0.00", "0.00%", "mm/dd/yyyy"]
    for r, row_data in enumerate(loan_rows, 2):
        for c, (val, fmt) in enumerate(zip(row_data, loan_fmts), 1):
            _body_cell(ws_loans, r, c, val, fmt)

    loan_first_notes = [
        (1, "002001 = loan number from your Epic statement"),
        (2, "2020 = this loan is for the 2020 grant"),
        (3, "Purchase = matches the Type in the Schedule sheet"),
        (4, "Purchase = the original loan taken to buy the shares"),
        (5, "2020 = year this loan was issued"),
        (6, "$92,500 = principal loan amount"),
        (7, "0.0095 = 0.95% annual interest rate"),
        (8, "Due July 15, 2029"),
    ]
    for col, note in loan_first_notes:
        ws_loans.cell(row=2, column=col).comment = Comment(note, "Sample")

    ws_loans.cell(row=3, column=4).comment = Comment(
        "Interest = a separate loan Epic issued for accumulated interest on the Purchase loan", "Sample"
    )

    # ---- Prices sheet ----
    ws_prices = wb.create_sheet("Prices")
    _write_headers(ws_prices, ["Date", "Price"])
    ws_prices.cell(row=1, column=1).comment = Comment(
        "One row per year. Epic announces the share price each March. "
        "The first row's price is used as the baseline for all cap-gains calculations.", "Sample"
    )
    ws_prices.cell(row=1, column=2).comment = Comment("Share price in dollars on that date.", "Sample")

    price_rows = [
        (date(2020, 3, 1), 1.85),
        (date(2021, 3, 1), 2.20),
        (date(2022, 3, 1), 2.75),
        (date(2023, 3, 1), 3.10),
        (date(2024, 3, 1), 3.60),
        (date(2025, 3, 1), 4.25),
        (date(2026, 3, 1), 5.10),
    ]
    for r, (d, p) in enumerate(price_rows, 2):
        _body_cell(ws_prices, r, 1, d, "mm/dd/yyyy")
        _body_cell(ws_prices, r, 2, p, "\\$#,##0.00")

    ws_prices.cell(row=2, column=1).comment = Comment(
        "March 1, 2020 = the day Epic announced the 2020 share price", "Sample"
    )
    ws_prices.cell(row=2, column=2).comment = Comment("$1.85 per share in 2020", "Sample")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Vesting_Sample.xlsx"},
    )


# ============================================================
# EXPORT
# ============================================================

_SCHED_HEADERS = ["Year", "Type", "Shares", "Price", "Vest Start", "Periods",
                   "Exercise Date", "DP Shares"]
_LOAN_HEADERS = ["Loan #", "Grant Year", "Grant Type", "Loan Type", "Loan Year",
                  "Amount", "Rate", "Due Date"]
_PRICE_HEADERS = ["Date", "Price"]
_EVENT_HEADERS = ["Date", "Grant Year", "Grant Type", "Event Type",
                  "Granted Shares", "Grant Price", "Exercise Price",
                  "Vested Shares", "Cum Shares", "Price Increase",
                  "Share Price", "Income", "Cum Income",
                  "Cap Gains", "Price Cap Gains", "Total Cap Gains", "Cum Cap Gains"]

_HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
_BODY_FONT = Font(name="Arial", size=10)
_ALT_FILL = PatternFill("solid", fgColor="FFE8E7FC")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFFFF")


def _write_headers(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT


def _body_cell(ws, row, col, val, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = _BODY_FONT
    c.fill = _ALT_FILL if row % 2 == 0 else _WHITE_FILL
    if fmt:
        c.number_format = fmt


@router.get("/export/excel")
def export_excel(
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    grants_db = db.query(Grant).filter(Grant.user_id == user.id).order_by(Grant.year).all()
    loans_db = db.query(Loan).filter(Loan.user_id == user.id).order_by(Loan.due_date).all()
    prices_db = db.query(Price).filter(Price.user_id == user.id).order_by(Price.effective_date).all()

    wb = openpyxl.Workbook()

    # -- Schedule sheet --
    ws_sched = wb.active
    ws_sched.title = "Schedule"
    _write_headers(ws_sched, _SCHED_HEADERS)
    for i, g in enumerate(grants_db, 2):
        _body_cell(ws_sched, i, 1, g.year)
        _body_cell(ws_sched, i, 2, g.type)
        _body_cell(ws_sched, i, 3, g.shares, "#,##0")
        _body_cell(ws_sched, i, 4, g.price, "\\$#,##0.00")
        _body_cell(ws_sched, i, 5, g.vest_start, "mm/dd/yyyy")
        _body_cell(ws_sched, i, 6, g.periods)
        _body_cell(ws_sched, i, 7, g.exercise_date, "mm/dd/yyyy")
        _body_cell(ws_sched, i, 8, g.dp_shares, "#,##0")

    # -- Loans sheet --
    ws_loans = wb.create_sheet("Loans")
    _write_headers(ws_loans, _LOAN_HEADERS)
    for i, ln in enumerate(loans_db, 2):
        _body_cell(ws_loans, i, 1, ln.loan_number)
        _body_cell(ws_loans, i, 2, ln.grant_year)
        _body_cell(ws_loans, i, 3, ln.grant_type)
        _body_cell(ws_loans, i, 4, ln.loan_type)
        _body_cell(ws_loans, i, 5, ln.loan_year)
        _body_cell(ws_loans, i, 6, ln.amount, "\\$#,##0.00")
        _body_cell(ws_loans, i, 7, ln.interest_rate, "0.00%")
        _body_cell(ws_loans, i, 8, ln.due_date, "mm/dd/yyyy")

    # -- Prices sheet --
    ws_prices = wb.create_sheet("Prices")
    _write_headers(ws_prices, _PRICE_HEADERS)
    for i, p in enumerate(prices_db, 2):
        _body_cell(ws_prices, i, 1, p.effective_date, "mm/dd/yyyy")
        _body_cell(ws_prices, i, 2, p.price, "\\$#,##0.00")

    # -- Events sheet (use existing writer for formulas) --
    ws_events = wb.create_sheet("Events")
    _write_headers(ws_events, _EVENT_HEADERS)

    # Generate events for formula-based writing
    grants_dicts = [{
        "year": g.year, "type": g.type, "shares": g.shares, "price": g.price,
        "vest_start": datetime.combine(g.vest_start, datetime.min.time()),
        "periods": g.periods,
        "exercise_date": datetime.combine(g.exercise_date, datetime.min.time()),
        "dp_shares": g.dp_shares or 0,
    } for g in grants_db]

    prices_dicts = [{"date": datetime.combine(p.effective_date, datetime.min.time()), "price": p.price} for p in prices_db]

    loans_dicts = [{
        "grant_yr": ln.grant_year, "grant_type": ln.grant_type,
        "loan_type": ln.loan_type, "loan_year": ln.loan_year,
        "amount": ln.amount, "interest_rate": ln.interest_rate,
        "due": datetime.combine(ln.due_date, datetime.min.time()),
        "loan_number": ln.loan_number,
    } for ln in loans_db]

    if grants_dicts or prices_dicts:
        events = generate_all_events(grants_dicts, prices_dicts, loans_dicts)
        initial_price = prices_dicts[0]["price"] if prices_dicts else 0
        timeline = compute_timeline(events, initial_price)

        # Save to temp file so write_events_to_excel can open it
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp_path = tmp.name
            wb.save(tmp_path)

        write_events_to_excel(tmp_path, timeline, prices_dicts)

        # Re-read the file with events written
        import os
        with open(tmp_path, "rb") as f:
            content = f.read()
        os.unlink(tmp_path)
    else:
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Vesting.xlsx"},
    )
