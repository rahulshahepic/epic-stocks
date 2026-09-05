"""Bounds on a caller-supplied .xlsx before a parser is allowed near it.

An .xlsx is a zip archive of XML. The 5 MB upload cap bounds what arrives on
the wire; it does not bound what that expands into. Deflate reaches roughly
1000:1 on repetitive XML, so a 5 MB upload can carry several gigabytes of
sheet data, and `openpyxl.load_workbook` builds all of it in memory before the
endpoint sees a single cell. The app container is capped at 512 MB, so one
request ends the process for everyone using it.

`read_only=True` is not the answer either: openpyxl streams worksheet rows in
that mode but still loads `sharedStrings.xml` whole, and a bomb hidden in the
shared strings reaches memory just the same.

So the archive is measured before it is parsed:

  * The zip directory is read first and the declared sizes are summed. This is
    the defence that does the work: it costs no decompression, and CPython's
    zipfile stops a member's read at its declared `file_size` (then fails the
    CRC), so an archive cannot hand any reader — this one or openpyxl — more
    bytes than its header admits to.
  * Then every member is decompressed in fixed-size chunks and discarded,
    counting bytes on the way out. Redundant against the above by CPython's
    current behaviour, and kept anyway: it is the check that does not depend on
    a zipfile implementation detail staying true. Memory stays at one chunk.

The caps are far above any workbook this app produces (a full export of the
largest holding a row quota permits is a few megabytes) and far below what a
bomb needs to matter.

Deliberately in scaffold and free of app imports: this is upload hardening,
like body_limit.py, not equity logic. Any new endpoint that hands a
caller-supplied archive to a parser needs the same treatment.
"""
import logging
import zipfile

logger = logging.getLogger(__name__)

# What the archive may expand to in total. A workbook holding the largest
# position the row quotas allow lands in single-digit megabytes.
MAX_TOTAL_UNCOMPRESSED = 64 * 1024 * 1024
# What any single member may expand to. Same ceiling: one sheet is allowed to
# be the whole workbook, but not more than the workbook.
MAX_MEMBER_UNCOMPRESSED = 64 * 1024 * 1024
# A workbook this app reads has a handful of sheets plus fixed parts. Thousands
# of members is a different kind of bomb — cheap per member, expensive in
# aggregate — and no legitimate upload here comes close.
MAX_MEMBERS = 512
# Read size when verifying. Bounds the memory a single member can cost.
_CHUNK = 64 * 1024


class WorkbookRejected(ValueError):
    """The archive is not something this server will hand to a parser."""


def check_workbook_bytes(raw: bytes) -> None:
    """Raise WorkbookRejected unless `raw` is a zip that stays inside the caps.

    Returns None on success. Callers parse `raw` afterwards; this only decides
    whether parsing is safe to attempt.
    """
    import io

    try:
        zf = zipfile.ZipFile(io.BytesIO(raw))
    except (zipfile.BadZipFile, OSError, EOFError, ValueError):
        # Not a zip at all, so there is nothing here that can expand into
        # anything. Say nothing and let the caller's own parse attempt fail:
        # each one already has a message for "this is not a spreadsheet", and
        # this guard exists to bound size, not to re-answer that question.
        return

    with zf:
        infos = zf.infolist()
        if len(infos) > MAX_MEMBERS:
            raise WorkbookRejected("spreadsheet has too many internal parts")

        # First pass: what the archive says about itself, which is also the
        # most any reader can get out of it — see the module docstring. Costs
        # no decompression at all.
        declared_total = 0
        for info in infos:
            if info.file_size > MAX_MEMBER_UNCOMPRESSED:
                raise WorkbookRejected("spreadsheet is too large once decompressed")
            declared_total += info.file_size
            if declared_total > MAX_TOTAL_UNCOMPRESSED:
                raise WorkbookRejected("spreadsheet is too large once decompressed")

        # Second pass: count what actually comes out, so the cap does not rest
        # on zipfile continuing to honour the declared size.
        total = 0
        for info in infos:
            if info.is_dir():
                continue
            try:
                with zf.open(info) as member:
                    while True:
                        chunk = member.read(_CHUNK)
                        if not chunk:
                            break
                        total += len(chunk)
                        if total > MAX_TOTAL_UNCOMPRESSED:
                            raise WorkbookRejected(
                                "spreadsheet is too large once decompressed"
                            )
            except WorkbookRejected:
                raise
            except (zipfile.BadZipFile, OSError, EOFError, ValueError) as exc:
                raise WorkbookRejected("spreadsheet file is corrupt") from exc


def load_workbook_safely(raw: bytes, **kwargs):
    """Measure `raw`, then open it with openpyxl.

    Every keyword is passed through to `openpyxl.load_workbook`, so callers
    keep their own read_only/data_only choices. Raises WorkbookRejected when
    the archive is over the caps or cannot be read as one.
    """
    import io

    import openpyxl

    check_workbook_bytes(raw)
    return openpyxl.load_workbook(io.BytesIO(raw), **kwargs)
