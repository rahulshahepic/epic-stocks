"""Tests for importing Epic's own files and the paste-out repair loop.

Fixtures in test_data/ are synthetic — the prices, rates and balances in them
are invented round numbers, not Epic's.
"""
import io
import json
import os
import re
import sys
from datetime import date, datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
import pytest


def _as_date(v) -> date:
    return v.date() if isinstance(v, datetime) else date.fromisoformat(str(v)[:10])

from app.epic_import import (build_prompt, build_skeleton, derive_draft, reconcile,
                             draft_from_payload, is_blocked, parse_share_csv,
                             parse_statement_lines, parse_statement_pdf,
                             to_wizard_payload, validate_draft)
from app.epic_import.rules import (DownPayment, attribute_loan, classify_row,
                                   down_payment_in_stock, is_vest_taxed,
                                   parse_loan_name, reconcile_down_payments)
from tests.conftest import register_user

DATA = os.path.join(os.path.dirname(__file__), "..", "..", "test_data")


def statement_lines():
    with open(os.path.join(DATA, "epic_loan_statement.txt")) as fh:
        return fh.read().split("\n")


def csv_bytes():
    with open(os.path.join(DATA, "epic_share_summary.csv"), "rb") as fh:
        return fh.read()


# A grant schedule matching the synthetic fixture. Deliberately has no 2023 Free
# template, so the fixture exercises the "shift the nearest one" path.
CONTENT = {
    "grant_templates": [
        {"year": 2020, "type": "Purchase", "vest_start": "2021-09-30", "periods": 5,
         "exercise_date": "2020-12-31", "default_catch_up": True, "show_dp_shares": True,
         "default_purchase_due_date": "2029-07-15", "default_tax_due_date": None},
        {"year": 2021, "type": "Purchase", "vest_start": "2022-09-30", "periods": 4,
         "exercise_date": "2021-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": "2030-07-15", "default_tax_due_date": None},
        {"year": 2021, "type": "Bonus", "vest_start": "2022-09-30", "periods": 3,
         "exercise_date": "2021-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": None, "default_tax_due_date": None},
        {"year": 2022, "type": "Purchase", "vest_start": "2023-09-30", "periods": 4,
         "exercise_date": "2022-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": "2031-06-30", "default_tax_due_date": None},
        {"year": 2022, "type": "Bonus", "vest_start": "2023-09-30", "periods": 3,
         "exercise_date": "2022-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": None, "default_tax_due_date": None},
        {"year": 2022, "type": "Free", "vest_start": "2027-09-30", "periods": 1,
         "exercise_date": "2022-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": None, "default_tax_due_date": "2031-06-30"},
        {"year": 2023, "type": "Bonus", "vest_start": "2024-09-30", "periods": 3,
         "exercise_date": "2023-12-31", "default_catch_up": False, "show_dp_shares": False,
         "default_purchase_due_date": None, "default_tax_due_date": None},
    ],
    "bonus_schedule_variants": [],
    "loan_rates": {
        "interest": {"2021": 0.02, "2022": 0.025, "2023": 0.04},
        "tax": {"Catch-Up": {"2021": 0.02, "2022": 0.03}, "Bonus": {"2023": 0.04}},
        "purchase_original": {"2020": {"rate": 0.01, "due_date": "2029-07-15"},
                              "2021": {"rate": 0.015, "due_date": "2030-07-15"},
                              "2022": {"rate": 0.03, "due_date": "2031-06-30"}},
    },
}


@pytest.fixture()
def skeleton():
    sk, findings = build_skeleton(CONTENT)
    assert [f for f in findings if f.severity == "error"] == []
    return sk


@pytest.fixture()
def parsed():
    statement, f1 = parse_statement_lines(statement_lines())
    rows, f2 = parse_share_csv(csv_bytes())
    return statement, rows, f1 + f2


@pytest.fixture()
def drafted(skeleton, parsed):
    statement, rows, pf = parsed
    draft, df = derive_draft(statement, rows, skeleton)
    findings = pf + df + validate_draft(draft, statement, rows, skeleton)
    return draft, findings


def grant(draft, year, gtype):
    return next(g for g in draft.grants if g.year == year and g.type == gtype)


def codes(findings, severity=None):
    return [f.code for f in findings if severity is None or f.severity == severity]


# ============================================================
# PARSING
# ============================================================

def test_statement_rows_and_totals_parse():
    statement, findings = parse_statement_lines(statement_lines())
    assert [f for f in findings if f.severity == "error"] == []
    assert len(statement.loans) == 9
    assert statement.statement_date.isoformat() == "2024-02-01"
    assert statement.subtotals == {2029: 545000.00, 2030: 1218000.00, 2031: 2032000.00}
    assert round(sum(l.balance for l in statement.loans), 2) == statement.printed_total


def test_wrapped_loan_name_is_reassembled():
    statement, _ = parse_statement_lines(statement_lines())
    wrapped = next(l for l in statement.loans if l.loan_number == "100008")
    assert wrapped.name == "2022 Bonus/2022 Grant - Interest Loan - 2023"


def test_boilerplate_is_not_read_as_a_loan_name():
    statement, _ = parse_statement_lines(statement_lines())
    assert all("share plan team" not in l.name for l in statement.loans)


@pytest.mark.parametrize("name,expected", [
    ("2018 Grant - Purchase Loan", ("Purchase", None, ["2018 Grant"])),
    ("2018 Grant - Interest Loan - 2020", ("Interest", 2020, ["2018 Grant"])),
    ("2020 Bonus - Tax Loan - 2021", ("Tax", 2021, ["2020 Bonus"])),
    ("2020 Bonus/2020 Grant - Interest Loan - 2024",
     ("Interest", 2024, ["2020 Bonus", "2020 Grant"])),
    ("not a loan name", (None, None, [])),
])
def test_loan_name_grammar(name, expected):
    assert parse_loan_name(name) == expected


def test_tax_loan_routes_to_the_zero_basis_grant():
    known = {2020: {"Purchase", "Catch-Up"}}
    assert attribute_loan(["2020 Grant"], "Tax", known)[:2] == (2020, "Catch-Up")
    assert attribute_loan(["2020 Grant"], "Interest", known)[:2] == (2020, "Purchase")
    assert attribute_loan(["2020 Grant"], "Tax", {2020: {"Purchase"}})[:2] == (2020, "Purchase")


def test_loan_covering_two_grants_goes_to_the_bonus_side():
    known = {2020: {"Purchase", "Bonus"}}
    for descriptors in (["2020 Bonus", "2020 Grant"], ["2020 Grant", "2020 Bonus"]):
        assert attribute_loan(descriptors, "Interest", known)[:2] == (2020, "Bonus")


def test_unused_categories_are_dropped():
    rows, findings = parse_share_csv(csv_bytes())
    assert "Other" not in [r.label for r in rows]
    assert len(rows) == 9 and findings == []


@pytest.mark.parametrize("label,expected", [
    ("2020 Purchased", (2020, "Purchase")), ("2020 Catch-up", (2020, "Catch-Up")),
    ("2021 Bonus Shares", (2021, "Bonus")), ("2023 Free", (2023, "Free")),
    # The developer label must not be swallowed by the plain bonus rule.
    ("2020 Developer Bonus Shares", (2020, "Developer Bonus Shares")),
    ("2021 developer bonus shares", (2021, "Developer Bonus Shares")),
    ("2019 Legacy Award Conversion", (None, None)),
])
def test_row_labels_map_to_grant_types(label, expected):
    assert classify_row(label) == expected


def test_developer_bonus_loans_are_attributed_to_their_own_grant():
    known = {2020: {"Purchase", "Bonus", "Developer Bonus Shares"}}
    ltype, lyear, descriptors = parse_loan_name("2020 Developer Bonus - Tax Loan - 2023")
    assert (ltype, lyear) == ("Tax", 2023)
    assert attribute_loan(descriptors, ltype, known)[:2] == (2020, "Developer Bonus Shares")


def test_unqualified_grant_tax_loan_prefers_catch_up_over_developer_bonus():
    """L3 ordering is unchanged: the developer bonus is only the last resort."""
    both = {2020: {"Purchase", "Catch-Up", "Developer Bonus Shares"}}
    assert attribute_loan(["2020 Grant"], "Tax", both)[:2] == (2020, "Catch-Up")
    only_dev = {2020: {"Purchase", "Developer Bonus Shares"}}
    assert attribute_loan(["2020 Grant"], "Tax", only_dev)[:2] == (2020, "Developer Bonus Shares")


# ============================================================
# SKELETON — structure the import must not invent
# ============================================================

def test_catch_up_rows_come_from_the_purchase_template(skeleton):
    catch_up = skeleton.template(2020, "Catch-Up")
    purchase = skeleton.template(2020, "Purchase")
    assert catch_up is not None
    assert (catch_up.vest_start, catch_up.periods) == (purchase.vest_start, purchase.periods)
    assert skeleton.template(2021, "Catch-Up") is None   # that year has no catch-up


def test_default_bonus_variant_sets_the_period_count():
    content = {**CONTENT, "bonus_schedule_variants": [
        {"grant_year": 2021, "grant_type": "Bonus", "variant_code": "C",
         "periods": 4, "label": "C", "is_default": True},
        {"grant_year": 2021, "grant_type": "Bonus", "variant_code": "A",
         "periods": 2, "label": "A", "is_default": False},
    ]}
    sk, _ = build_skeleton(content)
    assert sk.template(2021, "Bonus").periods == 4


def test_rates_are_indexed_by_kind_and_year(skeleton):
    assert skeleton.rate_for("Interest", "Purchase", 2022) == 0.025
    assert skeleton.rate_for("Tax", "Catch-Up", 2021) == 0.02
    assert skeleton.rate_for("Tax", "Bonus", 2023) == 0.04
    assert skeleton.rate_for("Interest", "Purchase", 1999) is None


# ============================================================
# DERIVING A DRAFT
# ============================================================

def test_schedule_comes_from_the_template_not_the_files(drafted):
    draft, _ = drafted
    g = grant(draft, 2021, "Purchase")
    assert g.vest_start.isoformat() == "2022-09-30"   # template, not inferred
    assert g.periods == 4
    assert g.exercise_date.isoformat() == "2021-12-31"


def test_missing_template_shifts_the_nearest_one_and_says_so(drafted):
    draft, findings = drafted
    free = grant(draft, 2023, "Free")
    assert free.vest_start.isoformat() == "2028-09-30"   # 2022 Free shifted one year
    assert free.periods == 1
    assert any(f.code == "S1" and "2023 Free" in f.subject for f in findings)


def test_shares_and_basis_come_from_the_csv(drafted):
    draft, _ = drafted
    g = grant(draft, 2021, "Purchase")
    assert g.shares == 200000
    assert g.price == 12.00


def test_catch_up_basis_is_recognised_as_accrued_at_vest(drafted):
    rows, _ = parse_share_csv(csv_bytes())
    row = next(r for r in rows if r.label == "2020 Catch-up")
    assert is_vest_taxed(row)[0] is True
    draft, _ = drafted
    assert grant(draft, 2020, "Catch-Up").price == 0.0


def test_unvested_shares_with_no_unvested_value_mean_zero_basis(drafted):
    rows, _ = parse_share_csv(csv_bytes())
    row = next(r for r in rows if r.label == "2023 Bonus Shares")
    assert is_vest_taxed(row) == (True, "shares are unvested but carry no unvested value")
    draft, _ = drafted
    assert grant(draft, 2023, "Bonus").price == 0.0


def test_loans_are_nested_under_the_grant_they_belong_to(drafted):
    draft, _ = drafted
    assert {l.loan_number for l in grant(draft, 2020, "Catch-Up").loans} == {"100003", "100004"}
    assert {l.loan_number for l in grant(draft, 2020, "Purchase").loans} == {"100001", "100002"}
    assert {l.loan_number for l in grant(draft, 2022, "Bonus").loans} == {"100008", "100009"}


def test_prices_come_from_purchase_grant_basis_dated_by_year(drafted):
    draft, _ = drafted
    assert [(p.effective_date.isoformat(), p.price) for p in draft.prices] == [
        ("2020-01-01", 10.00), ("2021-01-01", 12.00), ("2022-01-01", 15.00)]


def test_unmapped_category_is_reported_not_dropped_silently(drafted):
    _, findings = drafted
    assert any(f.code == "G1" and "Legacy Award" in f.subject for f in findings)


def test_the_fixture_reconciles_end_to_end(drafted):
    draft, findings = drafted
    assert [f.as_dict() for f in findings if f.severity == "error"] == []
    assert is_blocked(findings) is False


# ============================================================
# CHECKS
# ============================================================

def test_a_misread_row_blocks_the_import(skeleton):
    lines = [l.replace("$500,000.00", "$500,000.99") for l in statement_lines()]
    statement, f = parse_statement_lines(lines)
    rows, _ = parse_share_csv(csv_bytes())
    draft, df = derive_draft(statement, rows, skeleton)
    findings = f + df + validate_draft(draft, statement, rows, skeleton)
    assert "C1" in codes(findings, "error")
    assert "C2" in codes(findings, "error")
    assert is_blocked(findings) is True


def test_wrong_attribution_is_an_error_but_does_not_block(skeleton):
    """The documents disagreeing is worth stopping on, but it is the user's call."""
    statement, f = parse_statement_lines(statement_lines())
    rows, _ = parse_share_csv(csv_bytes())
    for sl in statement.loans:
        sl.name = sl.name.replace("Tax Loan", "Interest Loan")
    draft, df = derive_draft(statement, rows, skeleton)
    findings = f + df + validate_draft(draft, statement, rows, skeleton)
    assert "C3" in codes(findings, "error")
    assert "C4" in codes(findings, "error")
    assert is_blocked(findings) is False


def test_share_count_disagreement_is_reported(skeleton, parsed):
    statement, rows, _ = parsed
    draft, _ = derive_draft(statement, rows, skeleton)
    grant(draft, 2021, "Purchase").shares = 42
    assert "C8" in codes(validate_draft(draft, statement, rows, skeleton), "error")


def test_a_rate_off_the_record_is_flagged(skeleton, parsed):
    statement, rows, _ = parsed
    draft, _ = derive_draft(statement, rows, skeleton)
    grant(draft, 2020, "Purchase").loans[1].interest_rate = 0.99
    assert "C9" in codes(validate_draft(draft, statement, rows, skeleton), "warning")


def test_purchase_loan_rates_are_not_checked_because_they_get_refinanced(skeleton, parsed):
    statement, rows, _ = parsed
    draft, _ = derive_draft(statement, rows, skeleton)
    purchase = next(l for l in grant(draft, 2020, "Purchase").loans
                    if l.loan_type == "Purchase")
    purchase.interest_rate = 0.99
    assert "C9" not in codes(validate_draft(draft, statement, rows, skeleton))


def test_shares_remaining_mismatch_is_reported(skeleton, parsed):
    statement, rows, _ = parsed
    rows[0].shares_remaining += 1
    draft, _ = derive_draft(statement, rows, skeleton)
    assert "C6" in codes(validate_draft(draft, statement, rows, skeleton), "warning")


def test_csv_only_raises_no_unreconcilable_loan_errors(skeleton):
    rows, _ = parse_share_csv(csv_bytes())
    draft, df = derive_draft(None, rows, skeleton)
    findings = df + validate_draft(draft, None, rows, skeleton)
    assert [f.as_dict() for f in findings if f.severity == "error"] == []


# ============================================================
# THE RULE INVENTORY
# ============================================================
# The ids are the vocabulary for reporting an import bug, so a rule the code can
# report and the reference cannot explain is a bug in its own right. These pin
# RULES.md, the diagnostics legend and the code to each other.

MODULE = os.path.join(os.path.dirname(__file__), "..", "app", "epic_import")
RULE_ID = r"[A-Z]\d+"


def source_of(*names) -> str:
    return "\n".join(open(os.path.join(MODULE, n)).read() for n in names)


def documented_rules() -> set:
    text = open(os.path.join(MODULE, "RULES.md")).read()
    return set(re.findall(rf"^### `({RULE_ID})`", text, re.M))


def reported_rules() -> set:
    """Every id the importer can put in front of a user: findings from any module,
    plus the ids the diagnostics report tags differences with."""
    findings = set(re.findall(rf'Finding\("({RULE_ID})"',
                              source_of(*[f for f in os.listdir(MODULE) if f.endswith(".py")])))
    report = set(re.findall(rf'"({RULE_ID})"', source_of("reconcile.py")))
    return findings | report


def test_every_rule_the_import_can_report_is_documented():
    missing = reported_rules() - documented_rules()
    assert missing == set(), f"RULES.md does not explain {sorted(missing)}"


def test_the_reference_documents_no_rule_the_code_does_not_have():
    stale = documented_rules() - reported_rules()
    assert stale == set(), f"RULES.md explains {sorted(stale)}, which nothing emits"


def test_the_diagnostics_legend_explains_every_rule_it_can_print():
    from app.epic_import.reconcile import _FIELD_RULE, _RULE_HELP
    printable = set(_FIELD_RULE.values()) | set(
        re.findall(rf'Difference\("[a-z]+", [^)]*?"({RULE_ID})"', source_of("reconcile.py")))
    assert printable - set(_RULE_HELP) == set()


def test_the_legend_matches_what_the_rules_actually_do():
    """A legend that describes the opposite of the code is worse than none."""
    from app.epic_import.reconcile import _RULE_HELP
    assert "bonus side" in _RULE_HELP["L4"]
    assert attribute_loan(["2020 Bonus", "2020 Grant"], "Interest",
                          {2020: {"Purchase", "Bonus"}})[:2] == (2020, "Bonus")


def test_blocking_checks_are_documented_as_blocking():
    from app.epic_import.draft import BLOCKING_CHECKS
    text = open(os.path.join(MODULE, "RULES.md")).read()
    assert all(f"`{code}`" in text for code in BLOCKING_CHECKS)
    for code in BLOCKING_CHECKS:
        assert code in text.split("Only ")[1].split("block an import")[0]


# ============================================================
# DOWN PAYMENTS PAID IN STOCK (G8)
# ============================================================

def one_grant_files(shares, basis, loan, sold, rate=0.03):
    """A minimal share summary + statement for a single 2022 purchase grant.

    Small enough that the down payment arithmetic is the only thing moving:
    the loan is the grant's whole balance, so the statement's own totals and
    the CSV's loan balance follow from it.
    """
    interest = round(loan * rate, 2)
    header = ("Grant,Shares Granted,Shares Sold,Shares Remaining,83b Shares,"
              "Cost Basis of Shares,Loan Balance,Loan Due Year,Annual Interest Due")
    row = (f"2022 Purchased,{shares},{sold},{shares - sold},,{basis},{loan},2031,{interest}")
    lines = [
        "Stock Loan Statement - February 1, 2024",
        "900000001",
        "Test Employee",
        f"Total Principal Balance: ${loan:,.2f}",
        "Loan Loan Principal Interest Interest Loan Due",
        "Number Name Balance Rate (to Date) Date",
        f"100007 2022 Grant - Purchase Loan ${loan:,.2f} {rate * 100:.2f}% $100.00 6/30/2031",
        f"Subtotal ${loan:,.2f} 2031",
        f"Total ${loan:,.2f} $100.00",
    ]
    return f"{header}\n{row}\n".encode(), lines


def draft_from_files(skeleton, csv_raw, lines):
    statement, f1 = parse_statement_lines(lines)
    rows, f2 = parse_share_csv(csv_raw)
    draft, f3 = derive_draft(statement, rows, skeleton)
    return draft, f1 + f2 + f3 + validate_draft(draft, statement, rows, skeleton)


# 300,000 shares at 15.00; the down payment minimum of 20,000 is 1,334 whole
# shares (20,010.00), leaving a loan of 4,479,990.00.
STOCK_DP = dict(shares=300_000, basis=4_500_000.0, loan=4_479_990.0, sold=1_334)


def test_a_down_payment_in_whole_shares_is_read_from_the_loan(skeleton):
    draft, findings = draft_from_files(skeleton, *one_grant_files(**STOCK_DP))
    assert grant(draft, 2022, "Purchase").dp_shares == -1_334
    assert "G8" in codes(findings, "info")
    assert "G2" not in codes(findings)
    assert [f.as_dict() for f in findings if f.severity == "error"] == []


def test_the_down_payment_finding_names_the_grant_and_the_shares(skeleton):
    _, findings = draft_from_files(skeleton, *one_grant_files(**STOCK_DP))
    message = next(f.message for f in findings if f.code == "G8")
    assert "1,334 shares" in message and "2022 Purchase" in message


def test_shares_that_no_down_payment_accounts_for_are_still_reported_as_sold(skeleton):
    draft, findings = draft_from_files(skeleton, *one_grant_files(**{**STOCK_DP, "sold": 5_000}))
    assert grant(draft, 2022, "Purchase").dp_shares == 0
    assert "G2" in codes(findings, "info")
    assert "G8" not in codes(findings)


def test_a_loan_paid_down_since_is_not_read_as_a_down_payment(skeleton):
    # The gap is a whole number of shares (33,333 at 15.00) but far above the
    # policy minimum — a paid-down loan, not a down payment.
    draft, findings = draft_from_files(
        skeleton, *one_grant_files(shares=300_000, basis=4_500_000.0,
                                   loan=4_000_005.0, sold=33_333))
    assert grant(draft, 2022, "Purchase").dp_shares == 0
    assert "G2" in codes(findings, "info")


def test_a_down_payment_in_cash_leaves_no_shares_to_explain(skeleton):
    # 20,000.00 exactly: the minimum, but not a whole number of shares.
    draft, findings = draft_from_files(
        skeleton, *one_grant_files(shares=300_000, basis=4_500_000.0,
                                   loan=4_480_000.0, sold=0))
    assert grant(draft, 2022, "Purchase").dp_shares == 0
    assert "G8" not in codes(findings) and "G2" not in codes(findings)


def test_the_fixture_reports_its_sold_shares_because_they_reconcile_with_nothing(drafted):
    draft, findings = drafted
    assert all(g.dp_shares == 0 for g in draft.grants)
    assert any(f.code == "G2" and "10,000" in f.message for f in findings)


def test_a_supplied_draft_cannot_hand_back_more_shares_than_epic_reports_gone(skeleton):
    csv_raw, lines = one_grant_files(**STOCK_DP)
    statement, _ = parse_statement_lines(lines)
    rows, _ = parse_share_csv(csv_raw)
    payload = {"grants": [{"year": 2022, "type": "Purchase", "shares": 300_000,
                           "price": 15.0, "dp_shares": -50_000, "loans": []}],
               "prices": []}
    draft, _ = draft_from_payload(payload, skeleton)
    findings = validate_draft(draft, statement, rows, skeleton)
    assert "C11" in codes(findings, "warning")


def test_down_payment_shares_are_checked_against_the_loan_they_paid(skeleton):
    csv_raw, lines = one_grant_files(**STOCK_DP)
    statement, _ = parse_statement_lines(lines)
    rows, _ = parse_share_csv(csv_raw)
    payload = {"grants": [{"year": 2022, "type": "Purchase", "shares": 300_000,
                           "price": 15.0, "dp_shares": -1_000,
                           "loans": [{"loan_number": "100007", "loan_type": "Purchase",
                                      "loan_year": 2022, "amount": 4_479_990.0,
                                      "interest_rate": 0.03, "due_date": "2031-06-30"}]}],
               "prices": []}
    draft, _ = draft_from_payload(payload, skeleton)
    findings = validate_draft(draft, statement, rows, skeleton)
    assert any(f.code == "C11" and "2022 Purchase" == f.subject for f in findings)


def test_the_repair_prompt_carries_the_down_payment_policy(skeleton, drafted):
    draft, findings = drafted
    prompt = build_prompt(draft, findings, None, skeleton)
    assert "dp_shares" in prompt
    assert "10%" in prompt and "20,000" in prompt


@pytest.mark.parametrize("shares,basis,loan,expected", [
    # The minimum rounded up to whole shares — a down payment paid in stock.
    (300_000, 4_500_000.0, 4_479_990.0, 1_334),
    # 10% of a small purchase, below the cap, landing on whole shares.
    (10_000, 100_000.0, 90_000.0, 1_000),
    # Not a whole number of shares: paid in cash.
    (300_000, 4_500_000.0, 4_480_000.0, None),
    # Whole shares, but far more than the minimum: a paid-down loan.
    (300_000, 4_500_000.0, 4_000_005.0, None),
    # No gap at all.
    (300_000, 4_500_000.0, 4_500_000.0, None),
])
def test_down_payment_in_stock_reads_the_gap(shares, basis, loan, expected):
    dp = down_payment_in_stock(2022, "Purchase", shares, basis / shares, loan, 0.10, 20_000.0)
    assert (dp.shares if dp else None) == expected


def test_down_payments_are_only_applied_on_an_exact_unique_match():
    a = DownPayment(2023, "Purchase", 5_361, 19_996.53)
    b = DownPayment(2024, "Purchase", 4_717, 20_000.08)
    c = DownPayment(2025, "Purchase", 3_930, 20_003.70)
    chosen, outcome = reconcile_down_payments([a, b, c], 8_647)
    assert outcome == "reconciled" and [d.year for d in chosen] == [2024, 2025]
    assert reconcile_down_payments([a, b, c], 9_000)[1] == "none"
    assert reconcile_down_payments([], 8_647)[1] == "none"
    assert reconcile_down_payments([a, b, c], 0)[1] == "none"


def test_a_total_several_combinations_could_explain_is_left_alone():
    twin_a = DownPayment(2023, "Purchase", 1_000, 10_000.0)
    twin_b = DownPayment(2024, "Purchase", 1_000, 10_000.0)
    chosen, outcome = reconcile_down_payments([twin_a, twin_b], 1_000)
    assert outcome == "ambiguous" and chosen == []


# ============================================================
# A REPAIRED DRAFT COMING BACK
# ============================================================

def test_a_repaired_draft_is_read_back(skeleton, drafted):
    draft, _ = drafted
    payload = to_wizard_payload(draft)
    payload["grants"][0]["shares"] = 123456
    back, findings = draft_from_payload(payload, skeleton)
    assert [f for f in findings if f.severity == "error"] == []
    assert back.origin == "supplied"
    assert back.grants[0].shares == 123456
    assert len(back.all_loans) == len(draft.all_loans)


def test_the_same_checks_run_on_a_repaired_draft(skeleton, parsed):
    """A check that only works on our own output is not a check."""
    statement, rows, _ = parsed
    draft, _ = derive_draft(statement, rows, skeleton)
    payload = to_wizard_payload(draft)
    payload["grants"][0]["loans"] = []          # drop a grant's loans entirely
    back, _ = draft_from_payload(payload, skeleton)
    findings = validate_draft(back, statement, rows, skeleton)
    assert "C1" in codes(findings, "error")
    assert is_blocked(findings) is True


def test_a_repaired_draft_cannot_change_the_vesting_schedule(skeleton, drafted):
    draft, _ = drafted
    payload = to_wizard_payload(draft)
    payload["grants"][0]["periods"] = 99
    payload["grants"][0]["vest_start"] = "1999-01-01"
    back, findings = draft_from_payload(payload, skeleton)
    template = skeleton.template(back.grants[0].year, back.grants[0].type)
    assert back.grants[0].periods == template.periods
    assert back.grants[0].vest_start == template.vest_start
    assert codes(findings, "warning").count("C10") == 2


def test_unreadable_json_says_what_is_wrong(skeleton):
    _, findings = draft_from_payload({"nope": 1}, skeleton)
    assert findings[0].severity == "error"
    assert "grants" in findings[0].message


def test_a_bad_grant_is_reported_without_losing_the_good_ones(skeleton, drafted):
    draft, _ = drafted
    payload = to_wizard_payload(draft)
    payload["grants"].append({"year": 2024, "type": "Purchase", "shares": "lots"})
    back, findings = draft_from_payload(payload, skeleton)
    assert len(back.grants) == len(draft.grants)
    assert "R1" in codes(findings, "error")


# ============================================================
# THE PROMPT
# ============================================================

def test_the_prompt_carries_what_an_assistant_needs(skeleton, parsed, drafted):
    statement, rows, _ = parsed
    draft, findings = drafted
    prompt = build_prompt(draft, findings, statement, skeleton,
                          "\n".join(statement_lines()), csv_bytes().decode())

    assert "Return only the JSON object." in prompt
    assert "do not change these" in prompt.lower()
    assert "2022-09-30" in prompt                       # the fixed schedule
    assert "0.025" in prompt                            # the rates on record
    assert "100008" in prompt                           # the source statement text
    assert "2021 Bonus Shares" in prompt                # the source CSV text
    assert '"loan_type": "Purchase"' in prompt          # the output contract
    assert "Subtotal" in prompt


def test_the_prompt_lists_what_actually_failed(skeleton, parsed):
    statement, rows, _ = parsed
    for sl in statement.loans:
        sl.name = sl.name.replace("Tax Loan", "Interest Loan")
    draft, df = derive_draft(statement, rows, skeleton)
    findings = df + validate_draft(draft, statement, rows, skeleton)
    prompt = build_prompt(draft, findings, statement, skeleton)
    assert "[C3]" in prompt and "[C4]" in prompt


def test_a_clean_draft_still_produces_a_usable_prompt(skeleton, drafted):
    draft, _ = drafted
    prompt = build_prompt(draft, [], None, skeleton)
    assert "no failures" in prompt


# ============================================================
# API
# ============================================================

def make_pdf(lines: list[str]) -> bytes:
    """Smallest valid PDF that renders one text line per input line."""
    def esc(s):
        return s.replace("\\", r"\\").replace("(", r"\(").replace(")", r"\)")

    content = ("BT /F1 9 Tf 36 756 Td 11 TL\n"
               + "".join(f"({esc(l)}) Tj T*\n" for l in lines if l.strip()) + "ET")
    objs = [
        "<</Type/Catalog/Pages 2 0 R>>",
        "<</Type/Pages/Kids[3 0 R]/Count 1>>",
        "<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]"
        "/Resources<</Font<</F1 4 0 R>>>>/Contents 5 0 R>>",
        "<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>",
        f"<</Length {len(content)}>>\nstream\n{content}\nendstream",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for i, o in enumerate(objs, 1):
        offsets.append(len(out))
        out += f"{i} 0 obj\n{o}\nendobj\n".encode()
    start = len(out)
    out += f"xref\n0 {len(objs) + 1}\n0000000000 65535 f \n".encode()
    for off in offsets:
        out += f"{off:010d} 00000 n \n".encode()
    out += (f"trailer\n<</Size {len(objs) + 1}/Root 1 0 R>>\n"
            f"startxref\n{start}\n%%EOF\n").encode()
    return bytes(out)


def test_real_pdf_extraction_finds_the_same_rows():
    statement, findings = parse_statement_pdf(make_pdf(statement_lines()))
    assert [f for f in findings if f.severity == "error"] == []
    assert len(statement.loans) == 9
    assert round(sum(l.balance for l in statement.loans), 2) == 3795000.00


def upload_files(with_pdf=True):
    files = {"share_csv": ("shares.csv", csv_bytes(), "text/csv")}
    if with_pdf:
        files["statement_pdf"] = ("statement.pdf", make_pdf(statement_lines()),
                                  "application/pdf")
    return files


def analyze(client, **extra):
    resp = client.post("/api/epic-import/analyze", files=upload_files(), **extra)
    assert resp.status_code == 200, resp.text
    return resp.json()


def test_analyze_returns_a_draft_and_writes_nothing(client):
    register_user(client)
    body = analyze(client)
    assert body["origin"] == "parsed"
    assert body["summary"]["grants"] == 8
    assert body["summary"]["loans"] == 9
    assert body["summary"]["total_shares"] == 679000   # the unmapped category is not imported
    assert client.get("/api/grants").json() == []
    assert client.get("/api/loans").json() == []


def test_analyze_always_offers_a_prompt(client):
    register_user(client)
    assert "Return only the JSON object." in analyze(client)["prompt"]


def test_analyze_requires_at_least_one_file(client):
    register_user(client)
    assert client.post("/api/epic-import/analyze", files={}).status_code == 400


def test_a_misread_statement_blocks(client):
    register_user(client)
    broken = [l.replace("$500,000.00", "$500,000.99") for l in statement_lines()]
    resp = client.post("/api/epic-import/analyze", files={
        "share_csv": ("shares.csv", csv_bytes(), "text/csv"),
        "statement_pdf": ("s.pdf", make_pdf(broken), "application/pdf"),
    })
    body = resp.json()
    assert body["blocked"] is True
    assert "C1" in [f["code"] for f in body["findings"]]
    assert "[C1]" in body["prompt"]


def test_a_repaired_draft_can_be_pasted_back(client):
    register_user(client)
    first = analyze(client)
    payload = first["wizard_payload"]
    payload["grants"][0]["shares"] = 4321
    second = analyze(client, data={"revised_json": json.dumps(payload)})
    assert second["origin"] == "supplied"
    assert second["wizard_payload"]["grants"][0]["shares"] == 4321
    assert "C8" in [f["code"] for f in second["findings"]]


def test_a_repaired_draft_can_arrive_as_a_json_file(client):
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    files = upload_files()
    files["revised_draft"] = ("fixed.json", json.dumps(payload).encode(), "application/json")
    resp = client.post("/api/epic-import/analyze", files=files)
    assert resp.status_code == 200, resp.text
    assert resp.json()["origin"] == "supplied"


def test_a_repaired_draft_can_arrive_as_the_apps_own_workbook(client):
    register_user(client)
    analyze(client)
    client.post("/api/wizard/submit", json={
        **analyze(client)["wizard_payload"], "clear_existing": True,
        "generate_payoff_sales": False})
    export = client.get("/api/export/excel").content

    files = upload_files()
    files["revised_draft"] = ("filled.xlsx", export, "application/vnd.ms-excel")
    body = client.post("/api/epic-import/analyze", files=files).json()
    assert body["origin"] == "supplied"
    assert body["summary"]["grants"] == 8
    assert body["summary"]["loans"] == 9


def test_unparseable_paste_says_what_to_do(client):
    register_user(client)
    resp = client.post("/api/epic-import/analyze", files=upload_files(),
                       data={"revised_json": "Sure! Here is the JSON: {oops"})
    assert resp.status_code == 400
    assert "starting at '{'" in resp.json()["detail"]


def test_the_draft_can_be_submitted_through_the_wizard(client):
    """Acceptance goes through the wizard, so what is signed off is the position."""
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    resp = client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})
    assert resp.status_code == 201, resp.text
    assert len(client.get("/api/grants").json()) == 8
    assert len(client.get("/api/loans").json()) == 9
    assert len(client.get("/api/prices").json()) == 3


def test_analyze_requires_authentication(client):
    assert client.post("/api/epic-import/analyze", files=upload_files()).status_code == 401


# ============================================================
# DIAGNOSTICS
# ============================================================

def diff_files(client, **overrides):
    export = client.get("/api/export/excel")
    assert export.status_code == 200, export.text
    files = {"export_xlsx": ("export.xlsx", export.content,
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    files.update(upload_files())
    files.update(overrides)
    return files


def test_diff_finds_nothing_when_the_data_came_from_the_same_files(client):
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})
    report = client.post("/api/epic-import/diff", files=diff_files(client)).json()["report"]
    assert report["differences"] == []


def test_an_as_of_export_leaves_out_what_came_later(client):
    """Reconciling against a document dated months ago needs the data of that
    date. Nothing is versioned, so as-of is read off the dates each record has."""
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})

    everything = client.get("/api/export/excel")
    assert everything.status_code == 200
    early = client.get("/api/export/excel", params={"as_of": "2021-06-30"})
    assert early.status_code == 200
    assert "Vesting_2021-06-30.xlsx" in early.headers["content-disposition"]

    def counts(resp):
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        out = {n: wb[n].max_row - 1 for n in ("Schedule", "Loans", "Prices")}
        wb.close()
        return out

    now, then = counts(everything), counts(early)
    assert then["Schedule"] < now["Schedule"]
    assert then["Loans"] < now["Loans"]
    assert then["Prices"] < now["Prices"]

    # Everything kept is genuinely on or before the date.
    wb = openpyxl.load_workbook(io.BytesIO(early.content), data_only=True)
    cutoff = date(2021, 6, 30)
    for r in wb["Schedule"].iter_rows(min_row=2, values_only=True):
        assert _as_date(r[6]) <= cutoff, r          # exercise_date
    for r in wb["Loans"].iter_rows(min_row=2, values_only=True):
        assert int(r[4]) <= cutoff.year, r          # loan_year
    for r in wb["Prices"].iter_rows(min_row=2, values_only=True):
        assert _as_date(r[0]) <= cutoff, r          # effective_date
    wb.close()

    # And an as-of export still reads as a diff baseline.
    body = client.post("/api/epic-import/diff",
                       files=diff_files(client, export_xlsx=(
                           "export.xlsx", early.content,
                           "application/vnd.openxmlformats-officedocument."
                           "spreadsheetml.sheet"))).json()
    assert body["report"]["counts"]["baseline"] == "workbook"


def test_a_bad_as_of_date_is_rejected(client):
    register_user(client)
    assert client.get("/api/export/excel", params={"as_of": "last Tuesday"}).status_code == 400


def test_the_dashboard_holdings_report_works_as_a_baseline(client):
    """The dashboard export is a formatted position statement, not a dataset.
    It is readable, and what it leaves out is said once rather than reported as
    a difference on every row."""
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})

    report = client.get("/api/export/holdings-report", params={"as_of": "2026-02-01"})
    assert report.status_code == 200, report.text
    body = client.post("/api/epic-import/diff", files=diff_files(client, export_xlsx=(
        "holdings.xlsx", report.content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))).json()

    rep = body["report"]
    assert rep["counts"]["baseline"] == "holdings report"
    # Everything was actually read and matched, not quietly read as empty.
    assert rep["counts"]["existing_grants"] == rep["counts"]["imported_grants"] > 0
    assert rep["counts"]["existing_loans"] == rep["counts"]["imported_loans"] > 0
    assert [d for d in rep["differences"] if d["field"] == ""] == []
    # Nothing it cannot carry is reported as a per-row disagreement.
    omitted = {"periods", "vest_start", "dp_shares", "election_83b", "loan_number"}
    for d in rep["differences"]:
        if d["field"] in omitted:
            assert d["entity"] == "—" and d["severity"] == "info", d
    assert [d for d in rep["differences"] if d["entity"] == "price"] == []
    assert "prices" in {d["field"] for d in rep["differences"]}
    assert "holdings report" in body["markdown"]


def test_a_workbook_that_is_neither_export_is_refused(client):
    register_user(client)
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.title = "Something Else"
    wb.save(buf)
    resp = client.post("/api/epic-import/diff", files=diff_files(
        client, export_xlsx=("nope.xlsx", buf.getvalue(),
                             "application/vnd.openxmlformats-officedocument."
                             "spreadsheetml.sheet")))
    assert resp.status_code == 400
    assert "Holdings Report" in resp.json()["detail"]


def test_diff_names_the_rule_behind_each_difference(client):
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})
    g = next(x for x in client.get("/api/grants").json()
             if x["year"] == 2022 and x["type"] == "Purchase")
    client.put(f"/api/grants/{g['id']}", json={"shares": 999})

    report = client.post("/api/epic-import/diff", files=diff_files(client)).json()["report"]
    shares = [d for d in report["differences"] if d["field"] == "shares"]
    assert len(shares) == 1
    assert shares[0]["rule"] == "G2" and shares[0]["severity"] == "error"


def test_diff_writes_nothing(client):
    register_user(client)
    payload = analyze(client)["wizard_payload"]
    client.post("/api/wizard/submit", json={
        **payload, "clear_existing": True, "generate_payoff_sales": False})
    before = client.get("/api/grants").json()
    client.post("/api/epic-import/diff", files=diff_files(client))
    assert client.get("/api/grants").json() == before


def test_diff_markdown_download(client):
    register_user(client)
    resp = client.post("/api/epic-import/diff.md", files=diff_files(client))
    assert resp.status_code == 200
    assert "attachment" in resp.headers["content-disposition"]
    assert "# Epic import reconciliation" in resp.text


def test_diff_rejects_something_that_is_not_an_export(client):
    register_user(client)
    resp = client.post("/api/epic-import/diff", files=diff_files(
        client, export_xlsx=("notes.txt", b"not a workbook", "text/plain")))
    assert resp.status_code == 400


def test_analyze_returns_the_draft_in_the_shape_the_wizard_loads(client):
    """The wizard populates itself from grants/loans/prices; the draft arrives
    in that shape so it can be reviewed there rather than as a file."""
    register_user(client)
    prefill = analyze(client)["wizard_prefill"]
    assert len(prefill["grants"]) == 8
    assert len(prefill["loans"]) == 9
    assert len(prefill["prices"]) == 3
    assert all(g["id"] < 0 for g in prefill["grants"])       # never mistaken for saved rows
    purchase = next(g for g in prefill["grants"]
                    if g["year"] == 2021 and g["type"] == "Purchase")
    assert purchase["shares"] == 200000
    assert {l["loan_number"] for l in prefill["loans"] if l["grant_year"] == 2020
            and l["grant_type"] == "Catch-Up"} == {"100003", "100004"}


# ============================================================
# REGRESSIONS FOUND BY DRIVING THE REPAIR LOOP END TO END
# ============================================================

def with_extra_column(lines):
    """Epic adds a status column, so no statement row matches any more."""
    import re as _re
    return [_re.sub(r"( \d{1,2}/\d{1,2}/\d{4})$", r"\1 Active", l)
            if _re.match(r"^\d{6} ", l) else l for l in lines]


def test_a_repaired_draft_clears_the_parse_errors_it_fixes(client):
    """A supplied draft supersedes the parse. Leaving the parse errors standing
    means a perfect repair still reads as failing and the loop never converges."""
    register_user(client)
    good = analyze(client)["wizard_payload"]

    broken = {"share_csv": ("shares.csv", csv_bytes(), "text/csv"),
              "statement_pdf": ("s.pdf", make_pdf(with_extra_column(statement_lines())),
                                "application/pdf")}
    first = client.post("/api/epic-import/analyze", files=broken).json()
    assert first["blocked"] is True
    assert "L1" in [f["code"] for f in first["findings"]]

    fixed = client.post("/api/epic-import/analyze", files=broken,
                        data={"revised_json": json.dumps(good)}).json()
    assert fixed["blocked"] is False
    assert fixed["reconciles"] is True
    assert [f for f in fixed["findings"] if f["severity"] == "error"] == []
    # The parse complaint is kept as context, not dropped and not counted.
    stale = next(f for f in fixed["findings"] if f["code"] == "L1")
    assert stale["severity"] == "info"
    assert "before your correction" in stale["message"]


def test_a_renamed_csv_column_is_an_error_not_a_silent_zero(client):
    """Without the cost basis column every grant prices at zero, which turns
    every capital gain into ordinary income. That must never pass quietly."""
    register_user(client)
    renamed = csv_bytes().replace(b"Cost Basis of Shares", b"Total Cost Basis")
    body = client.post("/api/epic-import/analyze", files={
        "share_csv": ("shares.csv", renamed, "text/csv")}).json()

    assert body["blocked"] is True
    g0 = next(f for f in body["findings"] if f["code"] == "G0")
    assert g0["severity"] == "error"
    assert "priced at zero" in g0["message"]
    assert "Total Cost Basis" in g0["message"]     # names the headers it did find
    assert body["summary"]["grants"] == 0


def test_a_missing_check_column_only_costs_us_that_check():
    rows, findings = parse_share_csv(
        csv_bytes().replace(b"Annual Interest Due", b"Yearly Interest"))
    assert rows                                     # still usable
    assert [f.code for f in findings] == ["G0"]
    assert findings[0].severity == "warning"


def test_no_grants_is_said_once_not_once_per_loan(skeleton):
    """A CSV we cannot read leaves every loan homeless; saying so 54 times buries
    the one finding that explains why."""
    statement, _ = parse_statement_lines(statement_lines())
    draft, findings = derive_draft(statement, [], skeleton)
    homeless = [f for f in findings if f.code == "L3"]
    assert len(homeless) == 1
    assert "9 loans" in homeless[0].message


# ============================================================
# RECONCILING AGAINST DATA THAT WAS NOT ENTERED BY AN IMPORT
# ============================================================

def baseline_loan(**kw):
    base = {"loan_number": "", "grant_year": 2021, "grant_type": "Purchase",
            "loan_type": "Interest", "loan_year": 2022, "amount": 18000.0,
            "interest_rate": 0.025, "due_date": "2030-07-15",
            "refinances_loan_number": ""}
    return {**base, **kw}


def only(diffs, **match):
    return [d for d in diffs if all(getattr(d, k) == v for k, v in match.items())]


def test_a_loan_matches_on_its_grant_when_the_numbers_are_placeholders(skeleton, drafted):
    """Data entered through the wizard carries generated loan numbers, so matching
    on the number alone reports every loan as both missing and extra."""
    draft, _ = drafted
    report = reconcile(draft, [], [baseline_loan(loan_number="wiz-2021-I2022")], [])

    # Neither reported missing on the statement side nor extra on the user's.
    assert only(report.differences, key="100006", field="") == []
    assert only(report.differences, key="wiz-2021-I2022") == []
    swap = only(report.differences, entity="loan", field="loan_number")
    assert len(swap) == 1 and swap[0].severity == "info"
    assert report.counts["renumbered_loans"] == 1


def test_taking_epics_loan_numbers_is_one_line_not_one_row_each(skeleton, drafted):
    """Every wizard-entered loan swaps its placeholder for Epic's number. Listing
    that per loan buries the real differences under dozens of rows saying the
    import worked."""
    draft, _ = drafted
    mine = [baseline_loan(loan_number=f"wiz-{i}", grant_year=dg.year, grant_type=dg.type,
                          loan_type=dl.loan_type, loan_year=dl.loan_year,
                          amount=dl.amount, interest_rate=dl.interest_rate,
                          due_date=dl.due_date.isoformat())
            for i, (dg, dl) in enumerate(draft.all_loans)]
    report = reconcile(draft, [], mine, [])

    swap = only(report.differences, entity="loan", field="loan_number")
    assert len(swap) == 1
    assert report.counts["renumbered_loans"] == len(mine) > 1
    assert str(len(mine)) in swap[0].note


def test_a_purchase_balance_a_rounding_apart_is_not_an_error(skeleton, drafted):
    """Epic states the principal outstanding to the penny; a user types the round
    original. Calling that a bad read hides the reads that really are bad."""
    draft, _ = drafted
    dg, dl = next((dg, dl) for dg, dl in draft.all_loans if dl.loan_type == "Purchase")
    near = reconcile(draft, [], [baseline_loan(
        loan_number=dl.loan_number, grant_year=dg.year, grant_type=dg.type,
        loan_type="Purchase", loan_year=dl.loan_year, amount=dl.amount - 3.70,
        interest_rate=dl.interest_rate, due_date=dl.due_date.isoformat())], [])
    rows = only(near.differences, key=dl.loan_number, field="amount")
    assert len(rows) == 1 and rows[0].severity == "info"
    assert "rounded" in rows[0].note

    # A gap big enough to be the wrong row is still an error.
    far = reconcile(draft, [], [baseline_loan(
        loan_number=dl.loan_number, grant_year=dg.year, grant_type=dg.type,
        loan_type="Purchase", loan_year=dl.loan_year, amount=dl.amount * 0.9,
        interest_rate=dl.interest_rate, due_date=dl.due_date.isoformat())], [])
    rows = only(far.differences, key=dl.loan_number, field="amount")
    assert len(rows) == 1 and rows[0].severity == "error"


def test_a_loan_year_the_statement_never_carried_is_not_a_disagreement(skeleton, drafted):
    """"2020 Grant - Purchase Loan" is named the same however many times it has
    been refinanced, so the statement cannot say which year it dates from."""
    draft, _ = drafted
    dg, dl = next((dg, dl) for dg, dl in draft.all_loans if dl.loan_type == "Purchase")
    assert dl.year_on_statement is False
    report = reconcile(draft, [], [baseline_loan(
        loan_number=dl.loan_number, grant_year=dg.year, grant_type=dg.type,
        loan_type="Purchase", loan_year=dl.loan_year + 2, amount=dl.amount,
        interest_rate=dl.interest_rate, due_date=dl.due_date.isoformat())], [])
    rows = only(report.differences, key=dl.loan_number, field="loan_year")
    assert len(rows) == 1 and rows[0].severity == "info"

    # An interest loan is named "- 2022", so there the year is the statement's word.
    dg2, dl2 = next((dg, dl) for dg, dl in draft.all_loans if dl.loan_type == "Interest")
    assert dl2.year_on_statement is True


def test_a_loan_past_the_statements_reach_is_a_projection_not_a_gap(skeleton, drafted):
    """Loans are drawn each year. One dated after the newest year on the statement
    is the user projecting forward, exactly like a projected share price."""
    draft, _ = drafted
    reach = max(dl.loan_year for _, dl in draft.all_loans)
    report = reconcile(draft, [], [
        baseline_loan(loan_number="wiz-future", loan_year=reach + 1, amount=1234.0),
        baseline_loan(loan_number="wiz-past", loan_year=reach - 9, amount=99.0),
    ], [])

    later = only(report.differences, key="wiz-future", field="")
    assert len(later) == 1 and later[0].severity == "info"
    # No Epic number on it, so it is the user's own.
    assert "no Epic loan number" in later[0].note
    earlier = only(report.differences, key="wiz-past", field="")
    assert len(earlier) == 1 and earlier[0].severity == "warning"


def test_a_later_loan_carrying_an_epic_number_is_not_called_a_projection(skeleton, drafted):
    """A loan drawn after the statement was issued lands in the same branch as a
    projection. Epic's numbers are all digits, so one of those means Epic issued
    it — say which it is rather than assuming the user made it up."""
    draft, _ = drafted
    reach = max(dl.loan_year for _, dl in draft.all_loans)
    report = reconcile(draft, [], [baseline_loan(loan_number="900001",
                                                 loan_year=reach + 1, amount=1234.0)], [])
    row = only(report.differences, key="900001", field="")
    assert len(row) == 1 and row[0].severity == "info"
    assert "newer statement" in row[0].note
    assert "projection" not in row[0].note


def test_a_statement_loan_you_do_not_have_is_a_warning_not_an_error(skeleton, drafted):
    """The row is on the statement verbatim, so the rules did not invent it —
    the user's data is behind, which is what an import is for."""
    draft, _ = drafted
    report = reconcile(draft, [], [], [])
    missing = [d for d in report.differences if d.entity == "loan" and d.field == ""]
    assert missing and all(d.severity == "warning" for d in missing)
    assert all("importing would add it" in d.note for d in missing)


def test_a_refinanced_purchase_loan_matches_the_one_still_outstanding(skeleton, drafted):
    """The statement shows the loan currently owed; the app keeps the whole chain.
    Matching the original instead reports its old rate and due date as errors."""
    draft, _ = drafted
    chain = [
        baseline_loan(loan_number="wiz-2021-orig", loan_type="Purchase", loan_year=2021,
                      amount=1200000.0, interest_rate=0.0307, due_date="2025-07-15"),
        baseline_loan(loan_number="wiz-2021-refi", loan_type="Purchase", loan_year=2023,
                      amount=1200000.0, interest_rate=0.015, due_date="2030-07-15",
                      refinances_loan_number="wiz-2021-orig"),
    ]
    report = reconcile(draft, [], chain, [])

    assert only(report.differences, field="interest_rate") == []   # matched the tip
    superseded = [d for d in report.differences if d.key == "wiz-2021-orig"]
    assert len(superseded) == 1 and superseded[0].severity == "info"
    assert "refinance chain" in superseded[0].note


def test_a_loan_filed_against_a_different_grant_reads_as_one_row(skeleton, drafted):
    """Epic puts a loan covering two grants on the bonus side; a user may have
    filed it under the purchase grant. That is one disagreement, not two."""
    draft, _ = drafted
    mine = baseline_loan(loan_number="wiz-2022-I2023", grant_year=2022,
                         grant_type="Purchase", loan_type="Interest", loan_year=2023,
                         amount=7000.0, interest_rate=0.04, due_date="2031-06-30")
    report = reconcile(draft, [], [mine], [])

    assert only(report.differences, key="100008", field="") == []
    assert only(report.differences, key="wiz-2022-I2023") == []
    gt = only(report.differences, field="grant_type")
    assert len(gt) == 1
    assert (gt[0].imported, gt[0].existing) == ("Bonus", "Purchase")


def test_a_price_past_the_last_purchase_grant_is_a_projection_not_a_gap(skeleton, drafted):
    """Prices come from purchase grants, so later years are the user's own
    forecasts — calling them missing is a category error."""
    draft, _ = drafted
    prices = [{"effective_date": "2030-01-01", "price": 9.66},
              {"effective_date": "2019-01-01", "price": 8.0}]
    report = reconcile(draft, [], [], prices)

    later = only(report.differences, entity="price", key="2030")
    assert len(later) == 1 and later[0].severity == "info"
    assert "projection" in later[0].note
    earlier = only(report.differences, entity="price", key="2019")
    assert len(earlier) == 1 and earlier[0].severity == "warning"


def test_down_payment_shares_the_files_could_not_explain_are_context(skeleton, drafted):
    """Where rule G8 could read no down payment the draft carries 0, which is
    "could not tell" rather than a figure to disagree with."""
    draft, _ = drafted
    g = grant(draft, 2021, "Purchase")
    mine = [{"year": 2021, "type": "Purchase", "shares": g.shares, "price": g.price,
             "periods": g.periods, "vest_start": g.vest_start,
             "exercise_date": g.exercise_date, "dp_shares": -4717, "election_83b": False}]
    dp = only(reconcile(draft, mine, [], []).differences, field="dp_shares")
    assert len(dp) == 1 and dp[0].severity == "info"


def test_a_down_payment_read_off_the_loan_is_a_real_disagreement(skeleton):
    """Where G8 did read one, a difference is the import disagreeing with your
    record, reported against the rule that produced it."""
    draft, _ = draft_from_files(skeleton, *one_grant_files(**STOCK_DP))
    g = grant(draft, 2022, "Purchase")
    assert g.dp_shares == -1_334
    mine = [{"year": 2022, "type": "Purchase", "shares": g.shares, "price": g.price,
             "periods": g.periods, "vest_start": g.vest_start,
             "exercise_date": g.exercise_date, "dp_shares": -1_000,
             "election_83b": False}]
    dp = only(reconcile(draft, mine, [], []).differences, field="dp_shares")
    assert len(dp) == 1 and dp[0].severity == "warning" and dp[0].rule == "G8"
    assert "G8" in dp[0].note


def test_an_import_does_not_wipe_down_payment_shares(client):
    """A draft carrying 0 means the files could not say, so prefilling the wizard
    with it would erase a real figure the moment someone accepted."""
    register_user(client)
    client.post("/api/grants", json={
        "year": 2021, "type": "Purchase", "shares": 1, "price": 1.0,
        "vest_start": "2022-09-30", "periods": 4, "exercise_date": "2021-12-31",
        "dp_shares": -4717,
    })
    body = analyze(client)
    kept = next(g for g in body["wizard_prefill"]["grants"]
                if g["year"] == 2021 and g["type"] == "Purchase")
    assert kept["dp_shares"] == -4717


def test_a_bonus_basis_that_is_not_the_year_price_means_taxed_at_vest(skeleton, parsed):
    """A fully vested bonus grant looks the same either way; the year's share
    price is the only thing that separates a price paid from accumulated value."""
    statement, rows, _ = parsed
    draft, findings = derive_draft(statement, rows, skeleton)
    # The fixture's 2022 bonus carries 20.00/share against a 15.00 share price.
    assert grant(draft, 2022, "Bonus").price == 0.0
    note = next(f for f in findings if f.code == "G3" and "2022 Bonus" in f.subject)
    assert "not the 2022 share price" in note.message
    # One that does match the year's price is left alone.
    assert grant(draft, 2021, "Bonus").price == 12.00


def test_an_import_does_not_wipe_price_projections(client):
    """The wizard deletes prices its payload omits. A draft only covers years with
    a purchase grant, so forecasts past that would vanish on accept."""
    register_user(client)
    client.post("/api/prices", json={"effective_date": "2030-01-01", "price": 9.66})
    prefill = analyze(client)["wizard_prefill"]

    years = {p["effective_date"][:4] for p in prefill["prices"]}
    assert "2030" in years
    kept = next(p for p in prefill["prices"] if p["effective_date"].startswith("2030"))
    assert kept["id"] > 0          # the real row, so the wizard can preserve it


def test_an_import_does_not_wipe_grants_the_files_never_mention(client):
    register_user(client)
    client.post("/api/grants", json={
        "year": 2015, "type": "Bonus", "shares": 500, "price": 1.0,
        "vest_start": "2016-09-30", "periods": 3, "exercise_date": "2015-12-31"})
    prefill = analyze(client)["wizard_prefill"]
    assert any(g["year"] == 2015 and g["id"] > 0 for g in prefill["grants"])


# ── Current share price ──────────────────────────────────────────────────────
# The documents carry only prices Epic has already announced. Between
# announcements a position valued from them alone reads low, so the importer
# reports staleness and accepts today's price rather than imputing one.

def test_analyze_flags_a_stale_price(client):
    register_user(client)
    body = analyze(client)
    latest = max(p["effective_date"] for p in body["draft"]["prices"])
    assert latest[:4] < date.today().strftime("%Y")
    assert body["price_is_stale"] is True


def test_a_supplied_current_price_reaches_the_wizard(client):
    register_user(client)
    body = client.post("/api/epic-import/analyze", files=upload_files(),
                       data={"current_price": "42.75"}).json()

    today = date.today().isoformat()
    assert body["price_is_stale"] is False
    assert {"effective_date": today, "price": 42.75} in body["wizard_payload"]["prices"]
    # And through the prefill the wizard actually renders from.
    assert any(p["effective_date"] == today and p["price"] == 42.75
               for p in body["wizard_prefill"]["prices"])


def test_a_supplied_price_survives_into_saved_data(client):
    """What the importer offers must be what signing off actually stores."""
    register_user(client)
    body = client.post("/api/epic-import/analyze", files=upload_files(),
                       data={"current_price": "42.75"}).json()
    resp = client.post("/api/wizard/submit", json={
        **body["wizard_payload"], "clear_existing": True,
        "generate_payoff_sales": False})
    assert resp.status_code == 201, resp.text

    saved = client.get("/api/prices").json()
    today = date.today().isoformat()
    assert any(p["effective_date"] == today and p["price"] == 42.75 for p in saved)


def test_a_price_older_than_the_files_is_ignored(client):
    register_user(client)
    body = client.post("/api/epic-import/analyze", files=upload_files(),
                       data={"current_price": "0"}).json()
    assert len(body["draft"]["prices"]) == 3
