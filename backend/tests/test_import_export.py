"""Tests for Excel import/export API endpoints."""
import sys
import os
import io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from tests.conftest import register_user

FIXTURE = os.path.join(os.path.dirname(__file__), "..", "..", "test_data", "fixture.xlsx")


# ============================================================
# IMPORT
# ============================================================

def test_import_fixture(client):
    """Import test fixture → correct row counts (12 grants, 8 prices, 21 loans)."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        resp = client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 201
    data = resp.json()
    assert data["grants"] == 12
    assert data["prices"] == 8
    assert data["loans"] == 21


def test_import_populates_tables(client):
    """After import, CRUD endpoints return the imported data."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    grants = client.get("/api/grants").json()
    loans = client.get("/api/loans").json()
    prices = client.get("/api/prices").json()
    assert len(grants) == 12
    assert len(loans) == 21
    assert len(prices) == 8


def test_import_wipes_existing_data(client):
    """Import replaces existing data — doesn't append."""
    register_user(client)
    # Create some existing data
    client.post("/api/prices", json={"effective_date": "2020-01-01", "price": 99.99})
    prices_before = client.get("/api/prices").json()
    assert len(prices_before) == 1

    # Import overwrites
    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    prices_after = client.get("/api/prices").json()
    assert len(prices_after) == 8  # Not 9


def test_import_preserves_other_users(client, make_client):
    """Import wipes only the importing user's data, not other users'."""
    register_user(client, "a@test.com")

    # User A creates a price
    client.post("/api/prices", json={"effective_date": "2020-01-01", "price": 99.99})

    # User B imports
    with make_client("b@test.com") as client_b:
        with open(FIXTURE, "rb") as f:
            client_b.post(
                "/api/import/excel",
                files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

    # User A's data is untouched
    prices_a = client.get("/api/prices").json()
    assert len(prices_a) == 1


def test_import_events_match_known_values(client):
    """After importing fixture, /api/events produces the known-good totals."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    events = client.get("/api/events").json()
    real_events = [e for e in events if not e.get("is_projected")]
    assert len(real_events) == 89
    assert real_events[-1]["cum_shares"] == 558500
    # Projected liquidation event is injected at the end
    projected = [e for e in events if e.get("is_projected")]
    assert len(projected) == 1
    assert projected[0]["event_type"] == "Liquidation (projected)"
    assert projected[0]["cum_shares"] == 0
    assert projected[0]["gross_proceeds"] > 0


def test_import_rejects_non_excel(client):
    register_user(client)
    resp = client.post(
        "/api/import/excel",
        files={"file": ("data.csv", b"a,b,c\n1,2,3", "text/csv")},
    )
    assert resp.status_code == 400


# ============================================================
# EXPORT
# ============================================================

def test_export_empty(client):
    """Export with no data returns a valid xlsx with 4 sheets."""
    register_user(client)
    resp = client.get("/api/export/excel")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Schedule" in wb.sheetnames
    assert "Loans" in wb.sheetnames
    assert "Prices" in wb.sheetnames
    assert "Events" in wb.sheetnames
    wb.close()


def test_export_has_data(client):
    """Export after import contains the imported data."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    resp = client.get("/api/export/excel")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    # Count data rows (skip header)
    sched_rows = sum(1 for r in range(2, 100) if wb["Schedule"].cell(r, 1).value is not None)
    loan_rows = sum(1 for r in range(2, 100) if wb["Loans"].cell(r, 6).value is not None)
    price_rows = sum(1 for r in range(2, 30) if wb["Prices"].cell(r, 1).value is not None)
    assert sched_rows == 12
    assert loan_rows == 21
    assert price_rows == 8
    wb.close()


def test_export_roundtrip(client):
    """Import fixture → export → re-import exported file → same event count."""
    register_user(client)

    # Import fixture
    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    events_before = client.get("/api/events").json()

    # Export
    resp = client.get("/api/export/excel")
    exported = resp.content

    # Re-import the exported file
    resp2 = client.post(
        "/api/import/excel",
        files={"file": ("exported.xlsx", io.BytesIO(exported), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp2.status_code == 201

    events_after = client.get("/api/events").json()
    assert len(events_after) == len(events_before)
    assert events_after[-1]["cum_shares"] == events_before[-1]["cum_shares"]


def test_import_rejects_duplicate_grants_in_sheet(client):
    """Import with two rows having the same year+type in Schedule sheet returns 400."""
    register_user(client)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    headers = ["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"]
    ws.append(headers)
    row = [2020, "Purchase", 10000, 1.99, "2021-03-01", 5, "2020-12-31", 0]
    ws.append(row)
    ws.append(row)  # duplicate

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/import/excel",
        files={"file": ("dup.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "Duplicate" in resp.json()["detail"]


# ============================================================
# IMPORT BACKUPS
# ============================================================

def test_import_creates_backup(client):
    """After import, a backup of the replaced data is created."""
    register_user(client)
    # First import to populate data
    with open(FIXTURE, "rb") as f:
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    # Second import should create a backup of the first import's data
    with open(FIXTURE, "rb") as f:
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    backups = client.get("/api/import/backups").json()
    assert len(backups) >= 1
    assert backups[0]["grant_count"] == 12
    assert backups[0]["price_count"] == 8
    assert backups[0]["loan_count"] == 21


def test_import_backup_trimmed_to_three(client):
    """Backups are capped at 3 per user."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        data = f.read()
    for _ in range(5):
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    backups = client.get("/api/import/backups").json()
    assert len(backups) <= 3


def test_import_no_backup_when_empty(client):
    """No backup is created when there's nothing to back up."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    backups = client.get("/api/import/backups").json()
    # First import has no previous data so no backup
    assert len(backups) == 0


def test_restore_backup(client):
    """Restoring a backup brings back the pre-import data."""
    register_user(client)
    # Import fixture (12 grants)
    with open(FIXTURE, "rb") as f:
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    grants_after_first = client.get("/api/grants").json()
    assert len(grants_after_first) == 12

    # Build a minimal xlsx with 1 grant
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"])
    ws.append([2024, "Bonus", 5000, 0.00, "2024-01-01", 4, "2024-12-31", 0])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    client.post("/api/import/excel",
        files={"file": ("one.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    # Now only 1 grant
    assert len(client.get("/api/grants").json()) == 1

    # Restore the backup
    backups = client.get("/api/import/backups").json()
    assert len(backups) >= 1
    resp = client.post(f"/api/import/backups/{backups[0]['id']}/restore")
    assert resp.status_code == 200
    assert resp.json()["restored_grants"] == 12
    # Should be back to 12 grants
    assert len(client.get("/api/grants").json()) == 12


def test_restore_backup_wrong_user(client, make_client):
    """Cannot restore another user's backup."""
    register_user(client, "user1@example.com")
    with open(FIXTURE, "rb") as f:
        data = f.read()
    # User1 imports twice to create a backup
    for _ in range(2):
        client.post("/api/import/excel",
            files={"file": ("fixture.xlsx", io.BytesIO(data), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    backups = client.get("/api/import/backups").json()
    assert len(backups) >= 1
    # User2 tries to restore user1's backup
    with make_client("user2@example.com") as client2:
        resp = client2.post(f"/api/import/backups/{backups[0]['id']}/restore")
        assert resp.status_code == 404
