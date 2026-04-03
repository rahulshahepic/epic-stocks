"""Tests for template download, partial import, and import validation."""
import sys, os, io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from datetime import date
from tests.conftest import register_user

FIXTURE = os.path.join(os.path.dirname(__file__), "..", "..", "test_data", "fixture.xlsx")


def _make_xlsx(sheets: dict[str, list[list]]) -> bytes:
    """Build a minimal xlsx with given sheets. Each sheet is {name: [[row1], [row2], ...]}."""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        if first:
            ws.title = name
            first = False
        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# TEMPLATE DOWNLOAD
# ============================================================

def test_template_download(client):
    register_user(client)
    resp = client.get("/api/import/template")
    assert resp.status_code == 200
    assert "Vesting_Template" in resp.headers["content-disposition"]
    # Verify it's a valid xlsx with expected sheets
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Schedule" in wb.sheetnames
    assert "Loans" in wb.sheetnames
    assert "Prices" in wb.sheetnames
    # Headers should be present
    assert wb["Schedule"].cell(1, 1).value == "Year"
    assert wb["Loans"].cell(1, 1).value == "Loan #"
    assert wb["Prices"].cell(1, 1).value == "Date"
    # Example data in row 2
    assert wb["Schedule"].cell(2, 2).value == "Purchase"
    wb.close()


# ============================================================
# PARTIAL IMPORT
# ============================================================

def test_import_prices_only(client):
    """Upload with only Prices sheet — should import prices without touching grants/loans."""
    register_user(client)

    # First add a grant via API
    client.post("/api/grants", json={
        "year": 2020, "type": "Purchase", "shares": 100, "price": 5.0,
        "vest_start": "2020-01-01", "periods": 5, "exercise_date": "2025-01-01",
    })

    # Now import prices-only xlsx
    xlsx = _make_xlsx({"Prices": [
        ["Date", "Price"],
        [date(2020, 1, 1), 5.0],
        [date(2021, 1, 1), 8.0],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("prices.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201
    data = resp.json()
    assert data["prices"] == 2
    assert data["grants"] == 0
    assert data["loans"] == 0
    assert "Prices" in data["sheets_imported"]
    assert "Schedule" not in data["sheets_imported"]

    # Grant should still exist
    grants = client.get("/api/grants").json()
    assert len(grants) == 1


def test_import_schedule_only(client):
    """Upload with only Schedule sheet."""
    register_user(client)

    # Add a price first
    client.post("/api/prices", json={"effective_date": "2020-01-01", "price": 5.0})

    xlsx = _make_xlsx({"Schedule": [
        ["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"],
        [2020, "Purchase", 100, 5.0, date(2020, 1, 1), 5, date(2025, 1, 1), 0],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("grants.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201
    assert resp.json()["grants"] == 1
    assert resp.json()["prices"] == 0

    # Price should still exist
    prices = client.get("/api/prices").json()
    assert len(prices) == 1


def test_import_no_recognized_sheets(client):
    """File with no recognized sheets returns 400."""
    register_user(client)
    xlsx = _make_xlsx({"RandomSheet": [["foo", "bar"]]})
    resp = client.post("/api/import/excel",
        files={"file": ("bad.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "No recognized sheets" in resp.json()["detail"]


# ============================================================
# IMPORT VALIDATION
# ============================================================

def test_import_rejects_empty_grant_type(client):
    register_user(client)
    xlsx = _make_xlsx({"Schedule": [
        ["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"],
        [2020, "", 100, 5.0, date(2020, 1, 1), 5, date(2025, 1, 1), 0],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("bad.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "type cannot be empty" in resp.json()["detail"]

def test_import_rejects_negative_shares(client):
    register_user(client)
    xlsx = _make_xlsx({"Schedule": [
        ["Year", "Type", "Shares", "Price", "Vest Start", "Periods", "Exercise Date", "DP Shares"],
        [2020, "Purchase", -10, 5.0, date(2020, 1, 1), 5, date(2025, 1, 1), 0],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("bad.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "shares must be positive" in resp.json()["detail"]

def test_import_rejects_negative_loan_rate(client):
    register_user(client)
    xlsx = _make_xlsx({"Loans": [
        ["Loan #", "Grant Year", "Grant Type", "Loan Type", "Loan Year", "Amount", "Rate", "Due Date"],
        ["L1", 2020, "Purchase", "Interest", 2020, 5000, -0.05, date(2025, 1, 1)],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("bad.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "interest_rate" in resp.json()["detail"]

def test_import_rejects_negative_price(client):
    register_user(client)
    xlsx = _make_xlsx({"Prices": [
        ["Date", "Price"],
        [date(2020, 1, 1), -5.0],
    ]})
    resp = client.post("/api/import/excel",
        files={"file": ("bad.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
    assert "price must be positive" in resp.json()["detail"]


# ============================================================
# TEMPLATE FILL-AND-IMPORT
# ============================================================

def test_template_fill_and_import(client):
    """
    Download template → write real data into it → import → verify state and events.
    Catches regressions where the template structure diverges from what the importer expects.
    """
    register_user(client)

    # Download the template
    resp = client.get("/api/import/template")
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))

    # --- Schedule: overwrite example rows with real data (clear all example rows first) ---
    ws = wb["Schedule"]
    for col in range(1, 10):
        ws.cell(row=3, column=col).value = None  # clear Catch-Up example row
    ws.cell(row=2, column=1).value = 2022
    ws.cell(row=2, column=2).value = "Purchase"
    ws.cell(row=2, column=3).value = 8000
    ws.cell(row=2, column=4).value = 12.00
    ws.cell(row=2, column=5).value = date(2022, 3, 1)
    ws.cell(row=2, column=6).value = 4
    ws.cell(row=2, column=7).value = date(2032, 3, 1)
    ws.cell(row=2, column=8).value = 0

    # --- Prices: overwrite example row ---
    wp = wb["Prices"]
    wp.cell(row=2, column=1).value = date(2022, 1, 1)
    wp.cell(row=2, column=2).value = 12.00

    # --- Loans: clear example row so importer sees zero loans ---
    wl = wb["Loans"]
    for col in range(1, 9):
        wl.cell(row=2, column=col).value = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/import/excel",
        files={"file": ("filled_template.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201, resp.json()
    data = resp.json()
    assert data["grants"] == 1
    assert data["prices"] == 1
    assert data["loans"] == 0

    # Events computation must succeed and produce vesting events for the grant
    events = client.get("/api/events").json()
    assert isinstance(events, list)
    assert len(events) > 0


# ============================================================
# EXPORT IDEMPOTENCY
# ============================================================

def _data_snapshot(client):
    """Return sorted, id-stripped dicts for grants, loans, and prices."""
    def _pick(rows, keys):
        return sorted([{k: r[k] for k in keys} for r in rows], key=str)

    grants = _pick(
        client.get("/api/grants").json(),
        ["year", "type", "shares", "price", "vest_start", "periods", "exercise_date", "dp_shares"],
    )
    loans = _pick(
        client.get("/api/loans").json(),
        ["grant_year", "grant_type", "loan_type", "loan_year", "amount", "interest_rate", "due_date"],
    )
    prices = _pick(
        client.get("/api/prices").json(),
        ["effective_date", "price"],
    )
    return grants, loans, prices


def test_export_idempotent(client):
    """
    Import fixture → export → re-import the export → data rows are field-for-field identical.
    Verifies that the exporter and importer are fully concordant for real data.
    """
    register_user(client)

    with open(FIXTURE, "rb") as f:
        client.post(
            "/api/import/excel",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    grants_before, loans_before, prices_before = _data_snapshot(client)

    # Export then re-import
    exported = client.get("/api/export/excel").content
    resp = client.post(
        "/api/import/excel",
        files={"file": ("exported.xlsx", io.BytesIO(exported), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 201

    grants_after, loans_after, prices_after = _data_snapshot(client)

    assert grants_after == grants_before, "grants changed after export→import"
    assert loans_after == loans_before, "loans changed after export→import"
    assert prices_after == prices_before, "prices changed after export→import"
