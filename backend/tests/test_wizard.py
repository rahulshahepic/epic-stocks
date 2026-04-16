"""Tests for wizard endpoints: parse-file and submit."""
import sys
import os
import io
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from tests.conftest import register_user

FIXTURE = os.path.join(os.path.dirname(__file__), "..", "..", "test_data", "fixture.xlsx")


# ── parse-file ────────────────────────────────────────────────────────────────

def test_parse_file_fixture(client):
    """Parsing the fixture file returns grants and prices without errors."""
    register_user(client)
    with open(FIXTURE, "rb") as f:
        resp = client.post(
            "/api/wizard/parse-file",
            files={"file": ("fixture.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["grants"]) == 12
    assert len(data["prices"]) == 8
    # Structural fields present
    assert data["grants"][0]["year"] is not None
    assert data["grants"][0]["type"] is not None
    assert data["grants"][0]["periods"] is not None


def test_parse_file_template_no_numbers(client):
    """Parsing an xlsx with rows but missing share counts returns nulls, not errors."""
    register_user(client)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    # Header row
    ws.append(["year", "type", "shares", "price", "vest_start", "periods", "exercise_date", "dp_shares"])
    # Data row with structural info but no shares
    ws.append([2021, "Purchase", None, None, "2022-03-01", 4, "2021-12-31", None])

    ws2 = wb.create_sheet("Prices")
    ws2.append(["effective_date", "price"])
    ws2.append(["2021-12-31", None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/wizard/parse-file",
        files={"file": ("template.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["grants"]) == 1
    assert data["grants"][0]["year"] == 2021
    assert data["grants"][0]["periods"] == 4
    assert "shares" not in data["grants"][0]  # shares is personal data, not in template
    assert len(data["prices"]) == 1
    assert data["prices"][0]["price"] is None  # tolerant


def test_parse_file_empty_returns_empty(client):
    """An xlsx with no data rows returns empty lists."""
    register_user(client)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    ws.append(["year", "type", "shares"])  # header only

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/wizard/parse-file",
        files={"file": ("empty.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    assert resp.json()["grants"] == []


def test_parse_file_invalid_returns_422(client):
    """A non-Excel file returns 422."""
    register_user(client)
    resp = client.post(
        "/api/wizard/parse-file",
        files={"file": ("bad.xlsx", b"not an excel file", "application/octet-stream")},
    )
    assert resp.status_code == 422


def test_parse_file_requires_auth(client):
    """Unauthenticated parse-file request is rejected."""
    resp = client.post(
        "/api/wizard/parse-file",
        files={"file": ("x.xlsx", b"", "application/octet-stream")},
    )
    assert resp.status_code in (401, 403)


# ── submit ────────────────────────────────────────────────────────────────────

MINIMAL_PAYLOAD = {
    "prices": [{"effective_date": "2021-12-31", "price": 2.50}],
    "grants": [
        {
            "year": 2021,
            "type": "Purchase",
            "shares": 1000,
            "price": 2.50,
            "vest_start": "2022-03-01",
            "periods": 4,
            "exercise_date": "2021-12-31",
            "dp_shares": 0,
            "loans": [],
        }
    ],
}


def test_submit_creates_grant_and_price(client):
    """Wizard submit creates grants and prices."""
    register_user(client)
    resp = client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)
    assert resp.status_code == 201
    data = resp.json()
    assert data["grants"] == 1
    assert data["prices"] == 1
    assert data["loans"] == 0

    grants = client.get("/api/grants").json()
    assert len(grants) == 1
    assert grants[0]["year"] == 2021

    prices = client.get("/api/prices").json()
    assert len(prices) == 1


def test_submit_with_purchase_loan(client):
    """Wizard submit creates grant + Purchase loan."""
    register_user(client)
    payload = {
        "prices": [{"effective_date": "2021-12-31", "price": 2.50}],
        "grants": [
            {
                "year": 2021,
                "type": "Purchase",
                "shares": 1000,
                "price": 2.50,
                "vest_start": "2022-03-01",
                "periods": 4,
                "exercise_date": "2021-12-31",
                "dp_shares": 0,
                "loans": [
                    {
                        "loan_number": "111111",
                        "loan_type": "Purchase",
                        "loan_year": 2021,
                        "amount": 2000.00,
                        "interest_rate": 0.045,
                        "due_date": "2025-12-31",
                        "refinances_loan_number": "",
                    }
                ],
            }
        ],
    }
    resp = client.post("/api/wizard/submit", json=payload)
    assert resp.status_code == 201
    data = resp.json()
    assert data["loans"] == 1

    loans = client.get("/api/loans").json()
    assert len(loans) == 1
    assert loans[0]["loan_type"] == "Purchase"
    assert loans[0]["loan_number"] == "111111"


def test_submit_resolves_refinance_chain(client):
    """Loans referencing another by loan_number get refinances_loan_id resolved."""
    register_user(client)
    payload = {
        "prices": [{"effective_date": "2021-12-31", "price": 2.50}],
        "grants": [
            {
                "year": 2021,
                "type": "Purchase",
                "shares": 1000,
                "price": 2.50,
                "vest_start": "2022-03-01",
                "periods": 4,
                "exercise_date": "2021-12-31",
                "dp_shares": 0,
                "loans": [
                    {
                        "loan_number": "111111",
                        "loan_type": "Purchase",
                        "loan_year": 2021,
                        "amount": 2000.00,
                        "interest_rate": 0.045,
                        "due_date": "2025-12-31",
                        "refinances_loan_number": "",
                    },
                    {
                        "loan_number": "222222",
                        "loan_type": "Purchase",
                        "loan_year": 2023,
                        "amount": 2100.00,
                        "interest_rate": 0.05,
                        "due_date": "2028-12-31",
                        "refinances_loan_number": "111111",
                    },
                ],
            }
        ],
    }
    resp = client.post("/api/wizard/submit", json=payload)
    assert resp.status_code == 201
    assert resp.json()["loans"] == 2

    loans = client.get("/api/loans").json()
    original = next(l for l in loans if l["loan_number"] == "111111")
    refinance = next(l for l in loans if l["loan_number"] == "222222")
    assert refinance["refinances_loan_id"] == original["id"]

    # Payoff sale should only be generated for the active (non-refinanced) loan
    assert resp.json()["payoff_sales"] == 1
    sales = client.get("/api/sales").json()
    payoff_sales = [s for s in sales if s.get("loan_id")]
    assert len(payoff_sales) == 1
    assert payoff_sales[0]["loan_id"] == refinance["id"]


def test_submit_with_tax_loan(client):
    """Wizard submit creates a Catch-Up grant with a Tax loan."""
    register_user(client)
    payload = {
        "prices": [{"effective_date": "2021-12-31", "price": 2.50}],
        "grants": [
            {
                "year": 2021,
                "type": "Catch-Up",
                "shares": 500,
                "price": 0.0,
                "vest_start": "2022-03-01",
                "periods": 4,
                "exercise_date": "2021-12-31",
                "dp_shares": 0,
                "loans": [
                    {
                        "loan_number": "333333",
                        "loan_type": "Tax",
                        "loan_year": 2022,
                        "amount": 800.00,
                        "interest_rate": 0.05,
                        "due_date": "2026-12-31",
                        "refinances_loan_number": "",
                    }
                ],
            }
        ],
    }
    resp = client.post("/api/wizard/submit", json=payload)
    assert resp.status_code == 201
    data = resp.json()
    assert data["grants"] == 1
    assert data["loans"] == 1

    loans = client.get("/api/loans").json()
    assert loans[0]["loan_type"] == "Tax"
    assert loans[0]["grant_type"] == "Catch-Up"


def test_submit_clears_existing_data(client):
    """Submit with clear_existing=True replaces all prior data."""
    register_user(client)
    # Pre-populate
    client.post("/api/prices", json={"effective_date": "2020-01-01", "price": 1.50})
    assert len(client.get("/api/prices").json()) == 1

    resp = client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)
    assert resp.status_code == 201

    prices = client.get("/api/prices").json()
    assert len(prices) == 1
    assert prices[0]["price"] == 2.50  # replaced, not appended


def test_submit_multiple_grants(client):
    """Multiple grants across different years and types are all created."""
    register_user(client)
    payload = {
        "prices": [
            {"effective_date": "2020-12-31", "price": 2.00},
            {"effective_date": "2021-03-01", "price": 2.50},
        ],
        "grants": [
            {
                "year": 2020,
                "type": "Purchase",
                "shares": 800,
                "price": 2.00,
                "vest_start": "2021-03-01",
                "periods": 4,
                "exercise_date": "2020-12-31",
                "dp_shares": 0,
                "loans": [],
            },
            {
                "year": 2020,
                "type": "Catch-Up",
                "shares": 200,
                "price": 0.0,
                "vest_start": "2021-03-01",
                "periods": 4,
                "exercise_date": "2020-12-31",
                "dp_shares": 0,
                "loans": [],
            },
        ],
    }
    resp = client.post("/api/wizard/submit", json=payload)
    assert resp.status_code == 201
    data = resp.json()
    assert data["grants"] == 2
    assert data["prices"] == 2


def test_submit_requires_auth(client):
    """Unauthenticated submit is rejected."""
    resp = client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)
    assert resp.status_code in (401, 403)


# ── preview ───────────────────────────────────────────────────────────────────

def test_preview_empty_db_shows_all_added(client):
    """Preview with no existing data marks everything as 'added'."""
    register_user(client)
    resp = client.post("/api/wizard/preview", json=MINIMAL_PAYLOAD)
    assert resp.status_code == 200
    data = resp.json()
    assert all(g["status"] == "added" for g in data["grants"])
    assert all(p["status"] == "added" for p in data["prices"])


def test_preview_matching_data_shows_unchanged(client):
    """Preview with identical existing data marks everything 'unchanged'."""
    register_user(client)
    client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)

    resp = client.post("/api/wizard/preview", json=MINIMAL_PAYLOAD)
    assert resp.status_code == 200
    data = resp.json()
    assert all(g["status"] == "unchanged" for g in data["grants"])
    assert all(p["status"] == "unchanged" for p in data["prices"])


def test_preview_changed_shares_shows_updated(client):
    """Preview with different share count marks grant as 'updated'."""
    register_user(client)
    client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)

    modified = {**MINIMAL_PAYLOAD, "grants": [{**MINIMAL_PAYLOAD["grants"][0], "shares": 9999}]}
    resp = client.post("/api/wizard/preview", json=modified)
    assert resp.status_code == 200
    grants = resp.json()["grants"]
    assert grants[0]["status"] == "updated"
    assert grants[0]["shares"] == 9999
    assert grants[0]["old_shares"] == 1000


def test_preview_orphan_shows_removed(client):
    """Existing grant not in wizard payload is marked 'removed'."""
    register_user(client)
    # Submit two grants
    two_grant_payload = {
        "prices": MINIMAL_PAYLOAD["prices"],
        "grants": [
            MINIMAL_PAYLOAD["grants"][0],
            {**MINIMAL_PAYLOAD["grants"][0], "year": 2022, "type": "Bonus"},
        ],
    }
    client.post("/api/wizard/submit", json=two_grant_payload)

    # Preview with only the first grant
    resp = client.post("/api/wizard/preview", json=MINIMAL_PAYLOAD)
    assert resp.status_code == 200
    grants = resp.json()["grants"]
    statuses = {(g["year"], g["type"]): g["status"] for g in grants}
    assert statuses[(2021, "Purchase")] == "unchanged"
    assert statuses[(2022, "Bonus")] == "removed"


def test_preview_requires_auth(client):
    """Unauthenticated preview is rejected."""
    resp = client.post("/api/wizard/preview", json=MINIMAL_PAYLOAD)
    assert resp.status_code in (401, 403)


# ── merge mode (clear_existing=False) ────────────────────────────────────────

def test_merge_updates_existing_grant(client):
    """Merge mode updates share count on an existing grant without deleting others."""
    register_user(client)
    client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)

    modified = {**MINIMAL_PAYLOAD, "clear_existing": False, "grants": [{**MINIMAL_PAYLOAD["grants"][0], "shares": 2000}]}
    resp = client.post("/api/wizard/submit", json=modified)
    assert resp.status_code == 201

    grants = client.get("/api/grants").json()
    assert len(grants) == 1
    assert grants[0]["shares"] == 2000


def test_merge_deletes_orphaned_grant(client):
    """Merge mode removes existing grants not in the wizard payload."""
    register_user(client)
    two_grant_payload = {
        "prices": MINIMAL_PAYLOAD["prices"],
        "grants": [
            MINIMAL_PAYLOAD["grants"][0],
            {**MINIMAL_PAYLOAD["grants"][0], "year": 2022, "type": "Bonus"},
        ],
    }
    client.post("/api/wizard/submit", json=two_grant_payload)
    assert len(client.get("/api/grants").json()) == 2

    # Re-submit with only the 2021 grant
    resp = client.post("/api/wizard/submit", json={**MINIMAL_PAYLOAD, "clear_existing": False})
    assert resp.status_code == 201
    assert len(client.get("/api/grants").json()) == 1


def test_merge_preserves_grant_by_id(client):
    """preserve_grant_ids keeps an orphaned grant alive."""
    register_user(client)
    two_grant_payload = {
        "prices": MINIMAL_PAYLOAD["prices"],
        "grants": [
            MINIMAL_PAYLOAD["grants"][0],
            {**MINIMAL_PAYLOAD["grants"][0], "year": 2022, "type": "Bonus"},
        ],
    }
    client.post("/api/wizard/submit", json=two_grant_payload)
    grants = client.get("/api/grants").json()
    bonus_id = next(g["id"] for g in grants if g["year"] == 2022)

    # Re-submit with only 2021 grant but preserve the 2022 Bonus
    resp = client.post("/api/wizard/submit", json={
        **MINIMAL_PAYLOAD,
        "clear_existing": False,
        "preserve_grant_ids": [bonus_id],
    })
    assert resp.status_code == 201
    assert len(client.get("/api/grants").json()) == 2


def test_merge_keeps_manual_sales(client):
    """Merge mode does not delete user-created sales (only payoff sales)."""
    register_user(client)
    client.post("/api/wizard/submit", json=MINIMAL_PAYLOAD)
    # Add a manual sale (no loan_id)
    client.post("/api/sales", json={"date": "2023-01-01", "shares": 10, "price_per_share": 3.00, "notes": "manual"})
    sales_before = client.get("/api/sales").json()
    manual_sale = next(s for s in sales_before if s["notes"] == "manual")

    resp = client.post("/api/wizard/submit", json={**MINIMAL_PAYLOAD, "clear_existing": False})
    assert resp.status_code == 201
    sales_after = client.get("/api/sales").json()
    assert any(s["id"] == manual_sale["id"] for s in sales_after)
