"""
Excel I/O adapter for reading/writing Vesting.xlsx files.
Translates between spreadsheet rows and the plain dicts that core.py expects.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill


# ============================================================
# READER
# ============================================================

def read_grants_from_excel(ws):
    grants = []
    for i in range(2, 100):
        yr = ws.cell(row=i, column=1).value
        if yr is None:
            break
        grants.append({
            'year': int(yr),
            'type': ws.cell(row=i, column=2).value,
            'shares': int(ws.cell(row=i, column=3).value),
            'price': float(ws.cell(row=i, column=4).value),
            'vest_start': ws.cell(row=i, column=5).value,
            'periods': int(ws.cell(row=i, column=6).value),
            'exercise_date': ws.cell(row=i, column=14).value,
            'dp_shares': int(ws.cell(row=i, column=15).value or 0),
        })
    return grants


def read_prices_from_excel(ws):
    prices = []
    for i in range(2, 30):
        d = ws.cell(row=i, column=1).value
        p = ws.cell(row=i, column=2).value
        if d is None:
            break
        prices.append({'date': d, 'price': float(p)})
    return prices


def read_loans_from_excel(ws):
    loans = []
    for i in range(2, 100):
        amt = ws.cell(row=i, column=6).value
        if amt is None:
            break
        loans.append({
            'loan_number': ws.cell(row=i, column=1).value,
            'grant_yr': ws.cell(row=i, column=2).value,
            'grant_type': ws.cell(row=i, column=3).value,
            'loan_type': ws.cell(row=i, column=4).value.strip(),
            'loan_year': ws.cell(row=i, column=5).value,
            'amount': float(amt),
            'interest_rate': float(ws.cell(row=i, column=7).value),
            'due': ws.cell(row=i, column=8).value,
        })
    return loans


def read_all_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    grants = read_grants_from_excel(wb['Schedule'])
    prices = read_prices_from_excel(wb['Prices'])
    loans = read_loans_from_excel(wb['Loans'])
    initial_price = prices[0]['price'] if prices else 0
    wb.close()
    return grants, prices, loans, initial_price


# ============================================================
# WRITER
# ============================================================

_SCHED_ROW_OFFSET = 2
_PRICE_ROW_OFFSET = 2
_LOAN_ROW_OFFSET = 2


def write_events_to_excel(filepath, events, prices):
    """
    Write events to the Events sheet with Excel formulas
    referencing source sheets (Schedule, Loans, Prices).
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Events']

    for row in range(2, ws.max_row + 1):
        for col in range(1, 30):
            ws.cell(row=row, column=col).value = None
            ws.cell(row=row, column=col).fill = PatternFill()

    white = PatternFill('solid', fgColor='FFFFFFFF')
    alt = PatternFill('solid', fgColor='FFE8E7FC')
    font = Font(name='Arial', size=10)
    col_nf = {
        1: 'mm/dd/yyyy', 5: '#,##0', 6: '\\$#,##0.00', 7: '\\$#,##0.00',
        8: '#,##0', 9: '#,##0', 10: '\\$#,##0.00', 11: '\\$#,##0.00',
        12: '\\$#,##0.00', 13: '\\$#,##0.00', 14: '\\$#,##0.00',
        15: '\\$#,##0.00', 16: '\\$#,##0.00', 17: '\\$#,##0.00',
    }

    for idx, evt in enumerate(events):
        row = idx + 2
        fill = white if idx % 2 == 0 else alt
        src = evt.get('source') or {}
        src_type = src.get('type')

        def w(col, val):
            c = ws.cell(row=row, column=col)
            c.value = val
            c.fill = fill
            c.font = font
            c.number_format = col_nf.get(col, 'General')

        w(1, evt['date'])
        w(2, evt['grant_year'])
        w(3, evt['grant_type'])
        w(4, evt['event_type'])

        sr = src['index'] + _SCHED_ROW_OFFSET if src_type == 'grant' else None
        lr = src['index'] + _LOAN_ROW_OFFSET if src_type == 'loan' else None
        cr = src['index'] + _PRICE_ROW_OFFSET if src_type == 'price' else None
        pcr = src['prev_index'] + _PRICE_ROW_OFFSET if src_type == 'price' else None

        if evt['event_type'] == 'Exercise' and sr:
            w(5, f'=Schedule!C{sr}')
            w(6, f'=Schedule!D{sr}')
            w(7, f'=IF(Schedule!D{sr}=0,0,Schedule!D{sr})')
            w(8, None)
        elif evt['event_type'] == 'Down payment exchange' and sr:
            w(5, None); w(6, None); w(7, None)
            w(8, f'=Schedule!O{sr}')
        elif evt['event_type'] == 'Vesting' and sr:
            w(5, None)
            w(6, f'=Schedule!D{sr}')
            w(7, None)
            w(8, evt['vested_shares'])
        elif evt['event_type'] == 'Loan Repayment' and lr:
            w(5, None); w(6, None); w(7, None)
            w(8, f'=-ROUNDUP(Loans!F{lr}/K{row},0)')
        else:
            w(5, evt['granted_shares'])
            w(6, evt['grant_price'])
            w(7, evt['exercise_price'])
            w(8, evt['vested_shares'])

        w(9, f'=SUM(H$1:H{row})')

        if evt['event_type'] == 'Share Price' and cr and pcr:
            w(10, f'=Prices!B{cr}-Prices!B{pcr}')
        else:
            w(10, evt['price_increase'])

        first_price_row = _PRICE_ROW_OFFSET
        if row == 2:
            w(11, f'=Prices!B{first_price_row}')
        else:
            w(11, f'=K{row - 1}+J{row}')

        w(12, f'=IF(AND(H{row}>0,F{row}=0),H{row}*K{row},0)')
        w(13, f'=SUM(L$2:L{row})')
        w(14, f'=IF(AND(H{row}>0,F{row}>0),(K{row}-F{row})*H{row},0)')
        w(15, f'=J{row}*I{row}')
        w(16, f'=SUM(N{row}:O{row})')
        w(17, f'=SUM(P$2:P{row})')

    wb.save(filepath)
    return len(events)
